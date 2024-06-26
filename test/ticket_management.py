import os
import json
from datetime import datetime, timedelta, time
from collections import Counter, defaultdict
from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import (
    QMainWindow, QVBoxLayout, QWidget, QHBoxLayout, QPushButton, QLineEdit, QFileDialog, QTableWidget, QTableWidgetItem, QMessageBox,
    QTabWidget, QLabel, QListWidget, QStatusBar, QSplitter, QAbstractItemView, QListWidgetItem, QInputDialog, QApplication, QDialog, QTextEdit
)
from PyQt5.QtCore import QTimer, Qt, QRect, pyqtSlot
from PyQt5.QtGui import QFont, QIcon
from functools import partial
from ..beta_version.app.ui.dialogs import AdvancedFilterDialog, TopIncidentsDialog
from ..beta_version.app.detail.incidence_chart import TurnChart
from ..beta_version.app.ui.admin_dialog import AdminDialog
from ..beta_version.app.ui.excel_window import ExcelWindow

class TicketManagement(QMainWindow):
    def __init__(self, user):
        super().__init__()
        self.user = user
        self.excel_file = None
        self.config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "incidencias_config.json")
        self.mtbf_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mtbf_data.json")
        self.incidencias = AdminDialog.load_incidencias(self.config_file) or self.default_incidences()
        self.blocks = user.blocks if not user.is_admin else list(self.incidencias.keys())
        self.last_incidence_labels = {}
        self.incidences_count = {block: 0 for block in self.incidencias.keys()}
        self.pending_incidents = []
        self.filtered_incidents_data = {}
        self.incident_details = {block: Counter() for block in self.incidencias.keys()}
        self.mtbf_data = self.load_mtbf_data()
        self.state_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "incidence_state.json")
        self.mtbf_labels = {block: QLabel(f"MTBF {block}: N/A", self) for block in self.blocks}
        self.initUI()
        self.load_last_excel_file()
        self.load_incidence_state()
        self.reset_mtbf_timer()

    def default_incidences(self):
        return {
            "WC47 NACP": ["Etiquetadora", "Fallo en elevador", "No atornilla tapa", "Fallo tolva", "Fallo en paletizador", "No coge placa", "Palet atascado en la curva", "Ascensor no sube", "No pone tornillo", "Fallo tornillo", "AOI no detecta pieza", "No atornilla clips", "Fallo fijador tapa", "Secuencia atornillador", "Fallo atornillador", "Fallo cámara visión"],
            "WC48 P5F": ["Etiquetadora", "AOI (fallo etiqueta)", "AOI (malla)", "Cámara no detecta Pcb", "Cámara no detecta skeleton", "Cámara no detecta foams", "Cámara no detecta busbar", "Cámara no detecta foam derecho", "No detecta presencia power CP", "Tornillo atascado en tolva", "Cámara no detecta Power CP", "Cámara no detecta Top cover", "Detección de sealling mal puesto", "Robot no coge busbar", "Fallo etiqueta", "Power atascado en prensa, cuesta sacar", "No coloca bien el sealling"],
            "WC49 P5H": ["La cámara no detecta Busbar", "La cámara no detecta Top Cover", "Screw K30 no lo detecta puesto", "Atasco tuerca", "Tornillo atascado", "Etiquetadora", "Detección de sealling mal puesto", "No coloca bien el sealling", "Power atascado en prensa, cuesta sacar", "No lee QR"],
            "WV50 FILTER": ["Fallo cámara ferrite", "NOK Soldadura Plástico", "NOK Soldadura metal", "Traza", "NOK Soldad. Plástico+Metal", "Robot no coloca bien filter en palet", "No coloca bien la pcb", "QR desplazado", "Core enganchado", "Robot no coge PCB", "Fallo atornillador", "Pieza enganchada en HV Test", "Cover atascado", "Robot no coloca bien ferrita", "No coloca bien el core", "Fallo Funcional", "Fallo visión core", "Fallo cámara cover", "Repeat funcional", "Fallo cámara QR", "No coloca bien foam"],
            "SPL": ["Sensor de PCB detecta que hay placa cuando no la hay", "No detecta marcas Power", "Colisión placas", "Fallo dispensación glue", "Marco atascado en parte inferior", "Soldadura defectuosa", "Error en sensor de salida"]
        }

    def initUI(self):
        self.setWindowTitle(f"Ticket Management - {self.user.username}")
        self.resize(1200, 800)
        self.center_window_app()

        main_layout = QHBoxLayout()
        central_widget = QWidget()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

        self.splitter = QSplitter(Qt.Horizontal)
        main_layout.addWidget(self.splitter)

        self.tabWidget = QTabWidget(self)
        self.splitter.addWidget(self.tabWidget)

        for name in self.blocks:
            self.create_tab(name, self.incidencias[name])

        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        self.splitter.addWidget(right_widget)

        select_excel_button = QPushButton("Seleccionar Archivo Excel", self)
        select_excel_button.setStyleSheet(self.get_button_style())
        select_excel_button.clicked.connect(self.select_excel_file)
        right_layout.addWidget(select_excel_button)

        self.excel_path_display = QLineEdit(self)
        self.excel_path_display.setReadOnly(True)
        self.excel_path_display.setStyleSheet("background-color: #ffffff; border: 1px solid #cccccc; padding: 5px;")
        right_layout.addWidget(self.excel_path_display)

        self.view_excel_button = QPushButton("Ver Excel", self)
        self.view_excel_button.setStyleSheet(self.get_button_style())
        self.view_excel_button.clicked.connect(self.open_excel_window)
        right_layout.addWidget(self.view_excel_button)

        self.refresh_button = QPushButton("Refrescar", self)
        self.refresh_button.setStyleSheet(self.get_refresh_button_style())
        self.refresh_button.clicked.connect(self.update_all)
        right_layout.addWidget(self.refresh_button)

        self.turn_chart = TurnChart()  # Usamos la nueva clase para el gráfico por turnos

        self.daily_chart_button = QPushButton("Ver Gráfico Diario", self)
        self.daily_chart_button.setStyleSheet(self.get_button_style())
        self.daily_chart_button.clicked.connect(self.show_daily_chart)
        right_layout.addWidget(self.daily_chart_button)

        self.shift_chart_button = QPushButton("Ver Gráfico por Turno", self)
        self.shift_chart_button.setStyleSheet(self.get_button_style())
        self.shift_chart_button.clicked.connect(self.show_shift_chart)
        right_layout.addWidget(self.shift_chart_button)

        self.general_chart_button = QPushButton("Ver Gráfico General", self)
        self.general_chart_button.setStyleSheet(self.get_button_style())
        self.general_chart_button.clicked.connect(self.show_general_chart)
        right_layout.addWidget(self.general_chart_button)

        for block in self.user.blocks:
            mtbf_layout = QHBoxLayout()
            self.mtbf_labels[block].setStyleSheet("font-size: 16px; font-weight: bold; color: #4CAF50; background-color: #f0f0f0; padding: 5px; border-radius: 5px;")
            mtbf_layout.addWidget(self.mtbf_labels[block])

            info_button = QPushButton()
            info_button.setIcon(QIcon(os.path.join(os.path.dirname(__file__), "question_icon.png")))  # Asegúrate de que esta ruta sea correcta
            info_button.setToolTip("Haz clic para obtener más información sobre MTBF.")
            info_button.setStyleSheet("background-color: transparent; border: none; padding: 0px;")
            info_button.setFixedSize(24, 24)
            info_button.clicked.connect(self.show_mtbf_info)
            mtbf_layout.addWidget(info_button)

            right_layout.addLayout(mtbf_layout)

        self.global_incidence_list = QListWidget(self)
        self.global_incidence_list.setStyleSheet("QListWidget { background-color: #f0f0f0; border: 1px solid #ccc; padding: 5px; }")
        right_layout.addWidget(self.global_incidence_list)

        filter_button = QPushButton("Filtro Avanzado", self)
        filter_button.setStyleSheet(self.get_button_style())
        filter_button.clicked.connect(self.open_advanced_filter_dialog)
        right_layout.addWidget(filter_button)

        self.top_incidents_label = QLabel("Incidencias Más Relevantes")
        self.top_incidents_label.setFont(QFont("Arial", 14, QFont.Bold))
        self.top_incidents_label.setStyleSheet("color: #333333; margin: 10px 0;")
        right_layout.addWidget(self.top_incidents_label)

        self.view_details_button = QPushButton("Ver Detalles")
        self.view_details_button.setStyleSheet(self.get_button_style())
        self.view_details_button.clicked.connect(self.open_top_incidents_dialog)
        right_layout.addWidget(self.view_details_button)

        if self.user.is_admin:
            admin_button = QPushButton("Administrar Incidencias")
            admin_button.setStyleSheet(self.get_admin_button_style())
            admin_button.clicked.connect(self.open_admin_dialog)
            right_layout.addWidget(admin_button)

        right_layout.addStretch()

        self.status_bar = QStatusBar(self)
        self.setStatusBar(self.status_bar)

        self.detailed_messages_tab = QWidget()
        self.detailed_messages_layout = QVBoxLayout(self.detailed_messages_tab)
        self.tabWidget.addTab(self.detailed_messages_tab, "Mensajes detallados")
        self.detailed_messages_list = QListWidget()
        self.detailed_messages_list.setStyleSheet("QListWidget { background-color: #f0f0f0; border: 1px solid #ccc; padding: 5px; }")
        self.detailed_messages_layout.addWidget(self.detailed_messages_list)

        timer = QTimer(self)
        timer.timeout.connect(self.update_status_bar)
        timer.start(1000)

        self.apply_styles()
        self.update_top_incidents()
        self.update_mtbf_display()  # Asegúrate de actualizar la visualización del MTBF al iniciar la aplicación

    def center_window_app(self):
        screen_rect_app = QApplication.desktop().availableGeometry()
        window_width = self.width()
        window_height = self.height()
        x = (screen_rect_app.width() - window_width) // 2
        y = (screen_rect_app.height() - window_height) // 2
        self.setGeometry(QRect(x, y, window_width, window_height))

    def show_mtbf_info(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Información sobre MTBF")
        dialog.setFixedSize(400, 200)
        layout = QVBoxLayout()
        text_edit = QTextEdit()
        text_edit.setReadOnly(True)
        text_edit.setPlainText(
            "MTBF (Mean Time Between Failures) es una métrica que indica el tiempo promedio entre fallas en un sistema. "
            "Se calcula tomando el tiempo total de operación dividido por el número de fallas ocurridas durante ese tiempo. "
            "En esta aplicación, el MTBF se muestra en minutos y se actualiza cada vez que se registra una incidencia. "
            "El valor se reinicia cada 24 horas."
        )
        layout.addWidget(text_edit)
        close_button = QPushButton("Cerrar")
        close_button.setStyleSheet(self.get_button_style())
        close_button.clicked.connect(dialog.accept)
        layout.addWidget(close_button)
        dialog.setLayout(layout)
        dialog.exec_()

    def open_advanced_filter_dialog(self):
        self.filter_dialog = AdvancedFilterDialog(self)
        self.filter_dialog.exec_()

    def open_top_incidents_dialog(self):
        self.top_incidents_dialog = TopIncidentsDialog(self, incident_details=self.incident_details)
        self.top_incidents_dialog.exec_()

    def open_admin_dialog(self):
        self.admin_dialog = AdminDialog(self, incidencias=self.incidencias, config_file=self.config_file)
        self.admin_dialog.incidences_modified.connect(self.update_all)
        self.admin_dialog.exec_()

    def open_excel_window(self):
        if self.excel_file:
            self.excel_window = ExcelWindow(self.excel_file)
            self.excel_window.show()

    def update_all(self):
        self.update_top_incidents()
        self.update_global_incidence_list()
        self.update_tabs_incidences()
        self.update_mtbf_display()

    def update_tabs_incidences(self):
        for name in self.blocks:
            self.update_tab(name)

    def update_tab(self, block_name):
        for i in range(self.tabWidget.count()):
            tab = self.tabWidget.widget(i)
            if tab and self.tabWidget.tabText(i) == block_name:
                layout = tab.layout()
                list_widget = layout.itemAt(2).widget()
                list_widget.clear()
                for incidence in self.incidencias[block_name]:
                    item = QListWidgetItem(incidence)
                    list_widget.addItem(item)

    def update_global_incidence_list(self):
        current_fixing_incidents = []
        for i in range(self.global_incidence_list.count()):
            item = self.global_incidence_list.item(i)
            item_widget = self.global_incidence_list.itemWidget(item)
            if item_widget:
                labels = item_widget.findChildren(QLabel)
                if labels and (labels[1].text() == "Fixing" or labels[1].text() == "Pendiente"):
                    correct_button = item_widget.findChild(QPushButton, "correct_button")
                    details_button = item_widget.findChild(QPushButton, "details_button")
                    current_fixing_incidents.append((labels[0].text(), labels[1].text(), correct_button, details_button))

        self.global_incidence_list.clear()
        for incidence_text, fixing_label_text, correct_button, details_button in current_fixing_incidents:
            text_parts = incidence_text.split(" ")
            if len(text_parts) >= 5:
                date_str = text_parts[-3]
                time_str = text_parts[-2]
                block_name = " ".join(text_parts[:-5])

                item_widget = QWidget()
                item_layout = QVBoxLayout(item_widget)
                label_layout = QHBoxLayout()
                label_layout.addWidget(QLabel(incidence_text))
                item_layout.addLayout(label_layout)
                fixing_label = QLabel(fixing_label_text)
                fixing_label.setStyleSheet("color: red; font-weight: bold; font-size: 14px;" if fixing_label_text == "Fixing" else "color: orange; font-weight: bold; font-size: 14px;")
                correct_button.setParent(item_widget)
                correct_button.setStyleSheet(self.get_button_style())
                correct_button.setFixedSize(100, 30)
                details_button.setParent(item_widget)
                details_button.setStyleSheet(self.get_button_style())
                details_button.setFixedSize(150, 30)
                buttons_layout = QHBoxLayout()
                buttons_layout.addStretch()
                buttons_layout.addWidget(fixing_label)
                buttons_layout.addWidget(correct_button)
                buttons_layout.addWidget(details_button)
                item_layout.addLayout(buttons_layout)
                list_item = QListWidgetItem()
                list_item.setSizeHint(item_widget.sizeHint())
                self.global_incidence_list.addItem(list_item)
                self.global_incidence_list.setItemWidget(list_item, item_widget)
                correct_button.clicked.connect(partial(self.mark_incidence_as_fixed, block_name, incidence_text, date_str, time_str))
                details_button.clicked.connect(partial(self.add_incidence_details, block_name, incidence_text, date_str, time_str))

    def get_filtered_incidents(self, start_dt, end_dt, selected_block):
        if not self.excel_file or not os.path.exists(self.excel_file):
            return {}, {}

        workbook = load_workbook(self.excel_file)
        sheet = workbook.active

        filtered_counts = {block: {'count': 0, 'incidences': []} for block in self.incidencias.keys()}
        trends = {block: defaultdict(int) for block in self.incidencias.keys()}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            date_str, time_str, *incidences = row
            if date_str and time_str:
                try:
                    row_datetime = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
                    if start_dt <= row_datetime <= end_dt:
                        hour_str = row_datetime.strftime("%Y-%m-%d %H:00:00")
                        for i, incidence in enumerate(incidences):
                            if incidence != "-":
                                block_list = list(self.incidencias.keys())
                                if i < len(block_list):  # Verificar que i esté dentro del rango
                                    block = block_list[i]
                                    if selected_block == "Todos" or selected_block == block:
                                        filtered_counts[block]['count'] += 1
                                        filtered_counts[block]['incidences'].append(incidence)
                                        trends[block][hour_str] += 1
                except ValueError:
                    continue

        for block, data in filtered_counts.items():
            if data['incidences']:
                most_common_incidence, _ = Counter(data['incidences']).most_common(1)[0]
                data['most_common_incidence'] = most_common_incidence
            else:
                data['most_common_incidence'] = "N/A"

        self.filtered_incidents_data = filtered_counts

        if selected_block != "Todos":
            filtered_counts = {selected_block: filtered_counts[selected_block]}
            trends = {selected_block: trends[selected_block]}

        return filtered_counts, trends

    def get_filtered_incidents_by_date(self, date):
        start_dt = datetime.combine(date, time.min)
        end_dt = datetime.combine(date, time.max)
        return self.get_filtered_incidents(start_dt, end_dt, self.user.blocks[0])

    def get_filtered_incidents_by_shift(self, date, shift_start, shift_end):
        start_dt = datetime.combine(date, shift_start)
        end_dt = datetime.combine(date, shift_end)
        return self.get_filtered_incidents(start_dt, end_dt, self.user.blocks[0])

    def get_general_filtered_incidents(self, start_dt, end_dt):
        return self.get_filtered_incidents(start_dt, end_dt, "Todos")

    def create_tab(self, name, incidences):
        tab = QWidget()
        self.tabWidget.addTab(tab, name)
        layout = QVBoxLayout(tab)

        title = QLabel(name)
        title.setFont(QFont("Arial", 14, QFont.Bold))
        title.setStyleSheet("color: #333333; margin: 10px 0;")
        layout.addWidget(title)

        last_incidence_label = QLabel("Última incidencia: Ninguna")
        last_incidence_label.setFont(QFont("Arial", 12))
        last_incidence_label.setStyleSheet("color: #666666; margin-bottom: 10px;")
        layout.addWidget(last_incidence_label)
        self.last_incidence_labels[name] = last_incidence_label

        list_widget = QListWidget()
        font = QFont("Arial", 12)
        list_widget.setFont(font)
        list_widget.setStyleSheet("""
            QListWidget {
                background-color: #f9f9f9;
                border: 1px solid #ddd;
                padding: 10px;
            }
            QListWidget::item {
                padding: 10px;
                border-bottom: 1px solid #ddd;
            }
            QListWidget::item:selected {
                background-color: #4CAF50;
                color: white;
            }
        """)
        for incidence in incidences:
            item = QListWidgetItem(incidence)
            list_widget.addItem(item)

        layout.addWidget(list_widget)

        confirm_button = QPushButton("Confirmar Incidencia")
        confirm_button.setFont(QFont("Arial", 12))
        confirm_button.setFixedSize(500, 40)
        confirm_button.setStyleSheet(self.get_button_style())
        confirm_button.clicked.connect(lambda: self.confirm_incidence(name, list_widget))
        layout.addWidget(confirm_button)

    def confirm_incidence(self, block_name, list_widget):
        current_item = list_widget.currentItem()
        if current_item:
            incidence_text = current_item.text()
            timestamp = datetime.now()
            date_str = timestamp.strftime("%Y-%m-%d")
            time_str = timestamp.strftime("%H:%M:%S")
            self.last_incidence_labels[block_name].setText(f"Incidencia confirmada: {incidence_text} a las {time_str}")
            self.log_incidence_to_excel(block_name, date_str, time_str, incidence_text)

            self.update_mtbf(block_name, timestamp)

            QMessageBox.information(self, "Confirmación", "Incidencia confirmada.")

            fixing_label = QLabel("Fixing")
            fixing_label.setStyleSheet("color: red; font-weight: bold; font-size: 14px;")
            correct_button = QPushButton("Correct")
            correct_button.setObjectName("correct_button")
            correct_button.setStyleSheet(self.get_button_style())
            correct_button.setFixedSize(100, 30)
            details_button = QPushButton("Añadir Detalles")
            details_button.setObjectName("details_button")
            details_button.setStyleSheet(self.get_button_style())
            details_button.setFixedSize(150, 30)

            item_widget = QWidget()
            item_layout = QVBoxLayout(item_widget)
            label_layout = QHBoxLayout()
            label_layout.addWidget(QLabel(f"{block_name}: {incidence_text} a las {time_str} del {date_str}"))
            item_layout.addLayout(label_layout)

            buttons_layout = QHBoxLayout()
            buttons_layout.addStretch()
            buttons_layout.addWidget(fixing_label)
            buttons_layout.addWidget(correct_button)
            buttons_layout.addWidget(details_button)
            item_layout.addLayout(buttons_layout)

            list_item = QListWidgetItem()
            list_item.setSizeHint(item_widget.sizeHint())

            self.global_incidence_list.addItem(list_item)
            self.global_incidence_list.setItemWidget(list_item, item_widget)

            correct_button.clicked.connect(partial(self.mark_incidence_as_fixed, block_name, incidence_text, date_str, time_str))
            details_button.clicked.connect(partial(self.add_incidence_details, block_name, incidence_text, date_str, time_str))

            QTimer.singleShot(60000, partial(self.remind_user_to_fix, block_name, incidence_text, date_str, time_str, correct_button, details_button))

            self.update_top_incidents()
        else:
            QMessageBox.warning(self, "Ninguna Incidencia Seleccionada", "Selecciona una incidencia para confirmar.")

    def remind_user_to_fix(self, block_name, incidence_text, date_str, time_str, correct_button, details_button):
        for i in range(self.global_incidence_list.count()):
            item = self.global_incidence_list.item(i)
            item_widget = self.global_incidence_list.itemWidget(item)
            if item_widget:
                labels = item_widget.findChildren(QLabel)
                if labels and labels[0].text() == f"{block_name}: {incidence_text} a las {time_str} del {date_str}" and labels[1].text() == "Fixing":
                    msg_box = QMessageBox(self)
                    msg_box.setIcon(QMessageBox.Warning)
                    msg_box.setText(f'La incidencia "{incidence_text}" sigue en estado "Fixing".')
                    msg_box.setInformativeText("¿Deseas marcarla como 'Pendiente'?")
                    msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
                    yes_button = msg_box.button(QMessageBox.Yes)
                    yes_button.setText("Pendiente")
                    no_button = msg_box.button(QMessageBox.No)
                    no_button.setText("Fixing")
                    msg_box.exec_()

                    if msg_box.clickedButton() == yes_button:
                        labels[1].setText("Pendiente")
                        labels[1].setStyleSheet("color: orange; font-weight: bold; font-size: 14px;")
                        self.pending_incidents.append((block_name, incidence_text, date_str, time_str))
                        self.redirect_to_pending_incidence(block_name, incidence_text, labels[1], correct_button, details_button, date_str, time_str)
                    else:
                        QTimer.singleShot(60000, partial(self.remind_user_to_fix, block_name, incidence_text, date_str, time_str, correct_button, details_button))
                    break

    def redirect_to_pending_incidence(self, block_name, incidence_text, fixing_label, correct_button, details_button, date_str, time_str):
        for i in range(self.global_incidence_list.count()):
            item = self.global_incidence_list.item(i)
            item_widget = self.global_incidence_list.itemWidget(item)
            if item_widget:
                label = item_widget.findChild(QLabel)
                if label and incidence_text in label.text():
                    self.global_incidence_list.setCurrentRow(i)
                    self.global_incidence_list.scrollToItem(item, QAbstractItemView.PositionAtCenter)
                    break

    def mark_incidence_as_fixed(self, block_name, incidence_text, date_str, time_str):
        for i in range(self.global_incidence_list.count()):
            item = self.global_incidence_list.item(i)
            item_widget = self.global_incidence_list.itemWidget(item)
            if item_widget:
                labels = item_widget.findChildren(QLabel)
                if labels and labels[0].text() == f"{block_name}: {incidence_text} a las {time_str} del {date_str}":
                    fixing_label = labels[1]
                    fixing_label.setText("Reparada")
                    fixing_label.setStyleSheet("color: green; font-weight: bold; font-size: 14px;")
                    for btn in item_widget.findChildren(QPushButton):
                        if btn.objectName() == "correct_button":
                            btn.setEnabled(False)

                    repair_time_str = datetime.now().strftime("%H:%M:%S")
                    self.log_repair_time_to_excel(block_name, date_str, time_str, repair_time_str)
                    self.update_top_incidents()
                    break

    def add_incidence_details(self, block_name, incidence_text, date_str, time_str):
        detail_text, ok = QInputDialog.getMultiLineText(self, "Añadir Detalles", "Escribe los detalles de la incidencia:")
        if ok and detail_text:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            detail_message = f"{timestamp} - {block_name}: {incidence_text} ({date_str} {time_str})\nDetalles: {detail_text}"
            self.detailed_messages_list.addItem(detail_message)

    def log_repair_time_to_excel(self, block_name, date_str, time_str, repair_time_str):
        if self.excel_file and os.path.exists(self.excel_file):
            try:
                workbook = load_workbook(self.excel_file)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2):
                    if row[0].value == block_name and row[2].value == date_str and row[3].value == time_str:
                        repair_time_cell = sheet.cell(row=row[0].row, column=5)
                        repair_time_cell.value = repair_time_str

                        start_time = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
                        end_time = datetime.strptime(f"{date_str} {repair_time_str}", "%Y-%m-%d %H:%M:%S")
                        time_diff = end_time - start_time
                        time_diff_cell = sheet.cell(row=row[0].row, column=6)
                        time_diff_cell.value = str(time_diff)
                        break

                workbook.save(self.excel_file)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al registrar la hora de reparación en Excel: {e}")

    def select_excel_file(self):
        file_dialog = QFileDialog()
        file_dialog.setNameFilter("Archivos Excel (*.xlsx)")
        if file_dialog.exec_():
            self.excel_file = file_dialog.selectedFiles()[0]
            self.excel_path_display.setText(self.excel_file)
            with open(self.config_file, "w") as config:
                config.write(self.excel_file)
            self.update_top_incidents()

    def create_excel_if_not_exists(self, file_path):
        if not os.path.exists(file_path):
            workbook = Workbook()
            sheet = workbook.active
            headers = ["Bloque", "Incidencia", "Fecha", "Hora", "Hora de Reparación", "Tiempo de Reparación"]
            sheet.append(headers)
            workbook.save(file_path)


    def log_incidence_to_excel(self, block_name, date_str, time_str, incidence_text):
        if self.excel_file and os.path.exists(self.excel_file):
            try:
                workbook = load_workbook(self.excel_file)
                sheet = workbook.active

                headers = [cell.value for cell in sheet[1]]
                expected_headers = ["Bloque", "Incidencia", "Fecha", "Hora", "Hora de Reparación", "Tiempo de Reparación"]
                if headers != expected_headers:
                    sheet.delete_rows(1, 1)
                    sheet.insert_rows(1)
                    for idx, header in enumerate(expected_headers):
                        sheet.cell(row=1, column=idx + 1, value=header)

                new_row = [block_name, incidence_text, date_str, time_str, "", ""]
                sheet.append(new_row)

                workbook.save(self.excel_file)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al registrar la incidencia en Excel: {e}")


    def update_status_bar(self):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.statusBar().showMessage(f"Fecha y Hora Actual: {current_time}")

    def load_last_excel_file(self):
        if os.path.exists(self.config_file):
            with open(self.config_file, "r") as config:
                file_path = config.read().strip()
                if os.path.exists(file_path):
                    self.excel_file = file_path
                    self.excel_path_display.setText(file_path)
                    self.update_top_incidents()

    def update_top_incidents(self):
        self.incidences_count = {block: 0 for block in self.incidencias.keys()}
        self.incident_details = {block: Counter() for block in self.incidencias.keys()}

        if self.excel_file and os.path.exists(self.excel_file):
            workbook = load_workbook(self.excel_file)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= len(self.incidencias) + 4:
                    for i, block in enumerate(list(self.incidencias.keys()), start=2):
                        if row[i] and row[i] != "-":
                            self.incidences_count[block] += 1
                            self.incident_details[block][row[i]] += 1

    def show_daily_chart(self):
        today = datetime.today().date()
        incidents_daily, _ = self.get_filtered_incidents_by_date(today)
        self.turn_chart.plot_daily_chart(incidents_daily, "Incidencias Diarias")

    def show_shift_chart(self):
        today = datetime.today().date()
        shift_start = time(6, 0)
        shift_end = time(18, 0)
        incidents_shift, _ = self.get_filtered_incidents_by_shift(today, shift_start, shift_end)
        self.turn_chart.plot_shift_chart(incidents_shift, "Incidencias por Turno")

    def show_general_chart(self):
        start_dt = datetime.combine(datetime.today(), time.min)
        end_dt = datetime.combine(datetime.today(), time.max)
        incidents_general, _ = self.get_general_filtered_incidents(start_dt, end_dt)
        self.turn_chart.plot_general_chart(incidents_general, "Incidencias Generales")

    def apply_styles(self):
        title_font = QFont("Arial", 14, QFont.Bold)
        normal_font = QFont("Arial", 12)

        for i in range(self.tabWidget.count()):
            tab = self.tabWidget.widget(i)
            for widget in tab.findChildren(QLabel):
                if "Última incidencia" in widget.text():
                    widget.setFont(normal_font)
                else:
                    widget.setFont(title_font)

    def save_incidence_state(self):
        incidences = []
        for i in range(self.global_incidence_list.count()):
            item = self.global_incidence_list.item(i)
            item_widget = self.global_incidence_list.itemWidget(item)
            if item_widget:
                labels = item_widget.findChildren(QLabel)
                if labels:
                    incidence_text = labels[0].text()
                    status = labels[1].text() if len(labels) > 1 else "Fixing"
                    incidences.append({
                        "text": incidence_text,
                        "status": status
                    })
        state = {
            "user": self.user.username,
            "incidences": incidences
        }
        with open(self.state_file, "w") as file:
            json.dump(state, file, indent=4)

    def load_incidence_state(self):
        if os.path.exists(self.state_file):
            with open(self.state_file, "r") as file:
                state = json.load(file)
                if state["user"] == self.user.username:
                    for incidence in state["incidences"]:
                        incidence_text = incidence["text"]
                        status = incidence["status"]

                        text_parts = incidence_text.split(" ")
                        if len(text_parts) < 5:
                            continue

                        date_str = text_parts[-3]
                        time_str = text_parts[-2]
                        block_name = " ".join(text_parts[:-5])

                        fixing_label = QLabel(status)
                        if status == "Fixing":
                            fixing_label.setStyleSheet("color: red; font-weight: bold; font-size: 14px;")
                        elif status == "Pendiente":
                            fixing_label.setStyleSheet("color: orange; font-weight: bold; font-size: 14px;")
                        elif status == "Reparada":
                            fixing_label.setStyleSheet("color: green; font-weight: bold; font-size: 14px;")
                        correct_button = QPushButton("Correct")
                        correct_button.setObjectName("correct_button")
                        correct_button.setStyleSheet(self.get_button_style())
                        correct_button.setFixedSize(100, 30)
                        details_button = QPushButton("Añadir Detalles")
                        details_button.setObjectName("details_button")
                        details_button.setStyleSheet(self.get_button_style())
                        details_button.setFixedSize(150, 30)
                        item_widget = QWidget()
                        item_layout = QVBoxLayout(item_widget)
                        label_layout = QHBoxLayout()
                        label_layout.addWidget(QLabel(incidence_text))
                        item_layout.addLayout(label_layout)
                        buttons_layout = QHBoxLayout()
                        buttons_layout.addStretch()
                        buttons_layout.addWidget(fixing_label)
                        buttons_layout.addWidget(correct_button)
                        buttons_layout.addWidget(details_button)
                        item_layout.addLayout(buttons_layout)
                        list_item = QListWidgetItem()
                        list_item.setSizeHint(item_widget.sizeHint())
                        self.global_incidence_list.addItem(list_item)
                        self.global_incidence_list.setItemWidget(list_item, item_widget)
                        correct_button.clicked.connect(partial(self.mark_incidence_as_fixed, block_name, incidence_text, date_str, time_str))
                        details_button.clicked.connect(partial(self.add_incidence_details, block_name, incidence_text, date_str, time_str))
                        if status == "Fixing" or status == "Pendiente":
                            QTimer.singleShot(60000, partial(self.remind_user_to_fix, block_name, incidence_text, date_str, time_str, correct_button, details_button))

    def closeEvent(self, event):
        self.save_incidence_state()
        self.save_mtbf_data()
        event.accept()

    def update_mtbf(self, block_name, timestamp):
        if block_name in self.mtbf_data:
            mtbf_info = self.mtbf_data[block_name]
            if mtbf_info["last_time"] is not None:
                time_diff = (timestamp - mtbf_info["last_time"]).total_seconds() / 60.0  # Convertir a minutos
                mtbf_info["total_time"] += time_diff
                mtbf_info["incident_count"] += 1
            mtbf_info["last_time"] = timestamp
            self.update_mtbf_display()

    def update_mtbf_display(self):
        for block, data in self.mtbf_data.items():
            if block in self.mtbf_labels:
                if data["incident_count"] > 0:
                    mtbf = data["total_time"] / data["incident_count"]
                    self.mtbf_labels[block].setText(f"MTBF {block}: {mtbf:.2f} minutos")
                else:
                    self.mtbf_labels[block].setText(f"MTBF {block}: N/A")

    def reset_mtbf_timer(self):
        timer = QTimer(self)
        timer.timeout.connect(self.reset_mtbf_data)
        timer.start(24 * 60 * 60 * 1000)  # 24 horas en milisegundos

    def reset_mtbf_data(self):
        for block in self.mtbf_data.keys():
            self.mtbf_data[block] = {"last_time": None, "total_time": 0, "incident_count": 0}
        self.update_mtbf_display()

    def load_mtbf_data(self):
        if os.path.exists(self.mtbf_file):
            try:
                with open(self.mtbf_file, "r") as file:
                    data = json.load(file)
                    for block, mtbf_info in data.items():
                        if mtbf_info["last_time"] is not None:
                            mtbf_info["last_time"] = datetime.strptime(mtbf_info["last_time"], "%Y-%m-%d %H:%M:%S")
                    return data
            except json.JSONDecodeError:
                pass
        return {block: {"last_time": None, "total_time": 0, "incident_count": 0} for block in self.incidencias.keys()}

    def save_mtbf_data(self):
        with open(self.mtbf_file, "w") as file:
            data = {}
            for block, mtbf_info in self.mtbf_data.items():
                data[block] = mtbf_info.copy()
                if mtbf_info["last_time"] is not None:
                    data[block]["last_time"] = mtbf_info["last_time"].strftime("%Y-%m-%d %H:%M:%S")
            json.dump(data, file, indent=4)

    def get_button_style(self):
        return """
            QPushButton {
                background-color: #5C85FF;
                color: white;
                padding: 10px 20px;
                font-size: 16px;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #466BB7;
            }
        """

    def get_refresh_button_style(self):
        return """
            QPushButton {
                background-color: #FF8C42;
                color: white;
                padding: 10px 20px;
                font-size: 16px;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #CC7033;
            }
        """

    def get_admin_button_style(self):
        return """
            QPushButton {
                background-color: #FF5C5C;
                color: white;
                padding: 10px 20px;
                font-size: 16px;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #CC4A4A;
            }
        """