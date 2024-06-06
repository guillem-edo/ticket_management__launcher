import sys
import os
from datetime import datetime
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QTabWidget,
    QPushButton, QLabel, QListWidget, QListWidgetItem, QMessageBox, QLineEdit,
    QTableWidget, QTableWidgetItem, QStatusBar, QFileDialog, QDateEdit, QComboBox
)
from PyQt5.QtCore import QTimer, QDate, pyqtSignal, QRect
from PyQt5.QtGui import QFont
from openpyxl import Workbook, load_workbook
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

class LoginWindow(QWidget):
    login_successful = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Inicio de Sesión")
        self.resize(280, 360)
        self.center_window()

        layout = QVBoxLayout()
        logo_label = QLabel("Inicio de Sesión")
        layout.addWidget(logo_label)

        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("Usuario")
        layout.addWidget(self.username_input)

        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Contraseña")
        self.password_input.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.password_input)

        login_button = QPushButton("Confirmar")
        login_button.clicked.connect(self.login)
        layout.addWidget(login_button)

        self.setLayout(layout)

    def center_window(self):
        screen_rect = QApplication.desktop().availableGeometry()
        x = (screen_rect.width() - self.width()) // 2
        y = (screen_rect.height() - self.height()) // 2
        self.setGeometry(QRect(x, y, self.width(), self.height()))

    def login(self):
        username = self.username_input.text()
        password = self.password_input.text()
        if username == "pideu" and password == "1111":
            self.login_successful.emit()
            self.close()
        else:
            QMessageBox.warning(self, "Error de Inicio de Sesión", "Usuario o contraseña incorrectos")

class TicketManagement(QMainWindow):
    config_file = "config.txt"

    def __init__(self):
        super().__init__()
        self.initializeVariables()
        self.initUI()
        self.load_last_excel_file()

    def initializeVariables(self):
        self.excel_file = None
        self.blocks = ["WC47 NACP", "WC48 P5F", "WC49 P5H", "WV50 FILTER", "SPL"]
        self.last_incidence_labels = {}
        self.incidencias = {
            "WC47 NACP": ["Etiquetadora", "Fallo en elevador", "No atornilla tapa", "Fallo tolva",
                        "Fallo en paletizador", "No coge placa", "Palet atascado en la curva",
                        "Ascensor no sube", "No pone tornillo", "Fallo tornillo", "AOI no detecta pieza",
                        "No atornilla clips", "Fallo fijador tapa", "Secuencia atornillador",
                        "Fallo atornillador", "Fallo cámara visión"],
            "WC48 P5F": ["Etiquetadora", "AOI (fallo etiqueta)", "AOI (malla)", "Cámara no detecta Pcb", "Cámara no detecta skeleton",
                        "Cámara no detecta foams", "Cámara no detecta busbar", "Cámara no detecta foam derecho", "No detecta presencia power CP",
                        "Tornillo atascado en tolva", "Cámara no detecta Power CP", "Cámara no detecta Top cover", "Detección de sealling mal puesto",
                        "Robot no coge busbar", "Fallo etiqueta", "Power atascado en prensa, cuesta sacar", "No coloca bien el sealling"],
            "WC49 P5H": ["La cámara no detecta Busbar", "La cámara no detecta Top Cover", "Screw K30 no lo detecta puesto", "Atasco tuerca",
                        "Tornillo atascado", "Etiquetadora", "Detección de sealling mal puesto", "No coloca bien el sealling", "Power atascado en prensa, cuesta sacar",
                        "No lee QR"],
            "WV50 FILTER": ["Fallo cámara ferrite", "NOK Soldadura Plástico", "NOK Soldadura metal", "Traza", "NOK Soldad. Plástico+Metal", "Robot no coloca bien filter en palet",
                            "No coloca bien la pcb", "QR desplazado", "Core enganchado", "Robot no coge PCB", "Fallo atornillador", "Pieza enganchada en HV Test", "Cover atascado",
                            "Robot no coloca bien ferrita", "No coloca bien el core", "Fallo Funcional", "Fallo visión core", "Fallo cámara cover", "Repeat funcional", "Fallo cámara QR",
                            "No coloca bien foam"],
            "SPL": ["Sensor de PCB detecta que hay placa cuando no la hay", "No detecta marcas Power", "Colisión placas", "Fallo dispensación glue", "Marco atascado en parte inferior",
                    "Soldadura defectuosa", "Error en sensor de salida"]
        }
        self.incidences_count = {block: 0 for block in self.blocks}

    def initUI(self):
        self.setWindowTitle("Ticket Management")
        self.setGeometry(200, 200, 1200, 800)
        main_layout = QHBoxLayout()
        central_widget = QWidget()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)
        self.tabWidget = QTabWidget(self)
        main_layout.addWidget(self.tabWidget, 2)

        for name, incidences in self.incidencias.items():
            self.create_tab(name, incidences)

        right_layout = QVBoxLayout()
        select_excel_button = QPushButton("Seleccionar Archivo Excel", self)
        select_excel_button.clicked.connect(self.select_excel_file)
        right_layout.addWidget(select_excel_button)

        self.excel_path_display = QLineEdit(self)
        self.excel_path_display.setReadOnly(True)
        right_layout.addWidget(self.excel_path_display)

        self.table_widget = QTableWidget(self)
        self.table_widget.setFixedSize(600, 300)
        right_layout.addWidget(self.table_widget)

        self.global_incidence_list = QListWidget(self)
        right_layout.addWidget(self.global_incidence_list)

        self.figure = Figure()
        self.canvas = FigureCanvas(self.figure)
        right_layout.addWidget(self.canvas)

        self.date_from = QDateEdit(self)
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addDays(-30))
        right_layout.addWidget(self.date_from)

        self.date_to = QDateEdit(self)
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())
        right_layout.addWidget(self.date_to)

        self.block_selector = QComboBox(self)
        self.block_selector.addItems(["All"] + self.blocks)
        right_layout.addWidget(self.block_selector)

        filter_button = QPushButton("Filtrar", self)
        filter_button.clicked.connect(self.apply_filters)
        right_layout.addWidget(filter_button)

        right_layout.addStretch()
        right_widget = QWidget()
        right_widget.setLayout(right_layout)
        main_layout.addWidget(right_widget, 1)

        self.status_bar = QStatusBar(self)
        self.setStatusBar(self.status_bar)

        timer = QTimer(self)
        timer.timeout.connect(self.update_status_bar)
        timer.start(1000)

        self.apply_styles()

    def create_tab(self, name, incidences):
        tab = QWidget()
        self.tabWidget.addTab(tab, name)
        layout = QVBoxLayout(tab)
        title = QLabel(f"Incidencias - {name}")
        layout.addWidget(title)
        last_incidence_label = QLabel("Última incidencia: Ninguna")
        layout.addWidget(last_incidence_label)
        self.last_incidence_labels[name] = last_incidence_label
        list_widget = QListWidget()
        for incidence in incidences:
            item = QListWidgetItem(incidence)
            list_widget.addItem(item)
        layout.addWidget(list_widget)
        confirm_button = QPushButton("Confirmar Incidencia")
        confirm_button.clicked.connect(lambda: self.confirm_incidence(name, list_widget))
        layout.addWidget(confirm_button)

    def select_excel_file(self):
        file_dialog = QFileDialog()
        file_dialog.setNameFilter("Archivos Excel (*.xlsx)")
        if file_dialog.exec_():
            file_path = file_dialog.selectedFiles()[0]
            self.excel_file = file_path
            self.excel_path_display.setText(file_path)
            self.load_last_excel_file()
            self.update_graph()

    def load_last_excel_file(self):
        if os.path.exists(self.config_file):
            with open(self.config_file, "r") as config:
                file_path = config.read().strip()
                if os.path.exists(file_path):
                    self.excel_file = file_path
                    self.excel_path_display.setText(file_path)
                    self.update_excel_table()
                    self.update_graph()

    def update_excel_table(self, data=None):
        if self.excel_file and os.path.exists(self.excel_file):
            workbook = load_workbook(self.excel_file)
            sheet = workbook.active
            self.table_widget.setRowCount(sheet.max_row - 1)
            self.table_widget.setColumnCount(sheet.max_column)
            headers = [cell.value for cell in sheet[1]]
            self.table_widget.setHorizontalHeaderLabels(headers)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=0):
                for col_idx, value in enumerate(row):
                    item = QTableWidgetItem(str(value))
                    self.table_widget.setItem(row_idx, col_idx, item)

    def fetch_all_data(self):
        if self.excel_file and os.path.exists(self.excel_file):
            workbook = load_workbook(self.excel_file)
            sheet = workbook.active
            return [row for row in sheet.iter_rows(min_row=2, values_only=True)]
        return []

    def update_graph(self, data=None):
        if not data:
            data = self.fetch_all_data()

        self.incidences_count = {block: 0 for block in self.blocks}
        for row in data:
            for i, block in enumerate(self.blocks, start=2):
                if row[i] != "-":
                    self.incidences_count[block] += 1

        self.figure.clear()
        ax = self.figure.add_subplot(111)
        ax.bar(self.blocks, [self.incidences_count[block] for block in self.blocks])
        ax.set_xlabel('Bloques')
        ax.set_ylabel('Cantidad')
        ax.set_title('Total Incidencias')
        self.canvas.draw()

    def confirm_incidence(self, block_name, list_widget):
        current_item = list_widget.currentItem()
        if current_item:
            incidence_text = current_item.text()
            timestamp = datetime.now()
            date_str = timestamp.strftime("%Y-%m-%d")
            time_str = timestamp.strftime("%H:%M:%S")
            self.last_incidence_labels[block_name].setText(f"Incidencia confirmada: {incidence_text} a las {time_str}")
            self.log_incidence_to_excel(block_name, date_str, time_str, incidence_text)
            QMessageBox.information(self, "Confirmación", "Incidencia confirmada.")
            self.global_incidence_list.addItem(f"{block_name}: {incidence_text} a las {time_str} del {date_str}")
            self.update_excel_table()
            self.update_graph()

    def log_incidence_to_excel(self, block_name, date_str, time_str, incidence_text):
        if self.excel_file and os.path.exists(self.excel_file):
            try:
                workbook = load_workbook(self.excel_file)
                sheet = workbook.active

                if sheet.max_row == 1:  # Check if the header exists and is correct
                    headers = [cell.value for cell in sheet[1]]
                    expected_headers = ['Fecha', 'Hora'] + self.blocks
                    if headers != expected_headers:
                        sheet.insert_rows(1)
                        for i, header in enumerate(expected_headers):
                            sheet.cell(row=1, column=i+1, value=header)

                new_row_index = sheet.max_row + 1
                new_row = [date_str, time_str] + ['-' for _ in self.blocks]
                block_index = self.blocks.index(block_name) + 3  # +3 to adjust for Excel column start at 1 and include date and time
                new_row[block_index] = incidence_text
                for i, value in enumerate(new_row):
                    sheet.cell(row=new_row_index, column=i+1, value=value)
                workbook.save(self.excel_file)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Unable to log incidence to Excel: {e}")

    def apply_filters(self):
        selected_block = self.block_selector.currentText()
        date_from = self.date_from.date().toString("yyyy-MM-dd")
        date_to = self.date_to.date().toString("yyyy-MM-dd")
        filtered_data = self.filter_data(selected_block, date_from, date_to)
        self.update_graph(filtered_data)

    def filter_data(self, selected_block, date_from, date_to):
        if self.excel_file and os.path.exists(self.excel_file):
            workbook = load_workbook(self.excel_file)
            sheet = workbook.active
            filtered_incidences = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and date_from <= row[0] <= date_to:
                    block_index = self.blocks.index(selected_block) + 2 if selected_block != "All" else None
                    if block_index is None or (row[block_index] and row[block_index] != "-"):
                        filtered_incidences.append(row)
            return filtered_incidences

    def update_status_bar(self):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.statusBar().showMessage(f"Fecha y Hora Actual: {current_time}")

    def apply_styles(self):
        title_font = QFont("Arial", 14, QFont.Bold)
        normal_font = QFont("Arial", 12)
        for i in range(self.tabWidget.count()):
            tab = self.tabWidget.widget(i)
            for widget in tab.findChildren(QLabel):
                if "Última incidencia" in widget.text():
                    widget.setFont(normal_font)
                else:
                    widget.setFont(title_font)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    login_window = LoginWindow()
    ticket_management = TicketManagement()

    def on_login_successful():
        ticket_management.show()

    login_window.login_successful.connect(on_login_successful)
    login_window.show()

    sys.exit(app.exec_())