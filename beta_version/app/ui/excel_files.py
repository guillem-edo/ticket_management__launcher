import os 
import csv
from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from datetime import time, datetime

class excelDialogs():

    def __init__(self):
        self.config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.txt")
        print("excelDialogs instance created")
        
    def select_excel_file(self):
            file_dialog = QFileDialog()
            file_dialog.setNameFilter("Archivos Excel (*.xlsx)")
            if file_dialog.exec_():
                self.excel_file = file_dialog.selectedFiles()[0]
                self.excel_path_display.setText(self.excel_file)
                with open(self.config_file, "w") as config:
                    config.write(self.excel_file)
                self.update_top_incidents()

    def create_excel_if_not_exists(self, file_path):
        if not os.path.exists(file_path):
            workbook = Workbook()
            sheet = workbook.active
            headers = ["Bloque", "Incidencia", "Fecha", "Hora", "Turno", "Hora de Reparación", "Tiempo de Reparación", "MTBF"]
            sheet.append(headers)
            workbook.save(file_path)

    def load_last_excel_file(self):
        if os.path.exists(self.config_file):
            with open(self.config_file, "r") as config:
                file_path = config.read().strip()
                if os.path.exists(file_path):
                    self.excel_file = file_path
                    self.excel_path_display.setText(file_path)
                    self.update_top_incidents()
    
    def export_csv(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(self, "Guardar Informe CSV", "", "CSV Files (*.csv);;All Files (*)", options=options)
        if not file_name:
            return

        with open(file_name, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(["Bloque", "Número de Incidencias", "Incidencia más frecuente"])
            for row in range(self.results_table.rowCount()):
                bloque = self.results_table.item(row, 0).text() if self.results_table.item(row, 0) else ''
                num_incidencias = self.results_table.item(row, 1).text() if self.results_table.item(row, 1) else ''
                incidencia_frecuente = self.results_table.item(row, 2).text() if self.results_table.item(row, 2) else ''
                writer.writerow([bloque, num_incidencias, incidencia_frecuente])

            writer.writerow([])
            writer.writerow(["Incidencia", "Número de Incidencias"])
            for row in range(self.incidents_table.rowCount()):
                incidencia = self.incidents_table.item(row, 0).text() if self.incidents_table.item(row, 0) else ''
                num_incidencias = self.incidents_table.item(row, 1).text() if self.incidents_table.item(row, 1) else ''
                writer.writerow([incidencia, num_incidencias])

        QMessageBox.information(self, "Exportar Informe", "Informe exportado con éxito en formato CSV.")

    def log_repair_time_to_excel(self, block_name, date_str, time_str, repair_time_str):
        if self.excel_file and os.path.exists(self.excel_file):
            try:
                workbook = load_workbook(self.excel_file)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=False):
                    if row[0].value == block_name and row[2].value == date_str and row[3].value == time_str:
                        repair_time_cell = row[5]
                        repair_time_cell.value = repair_time_str

                        start_time = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
                        end_time = datetime.strptime(f"{date_str} {repair_time_str}", "%Y-%m-%d %H:%M:%S")
                        time_diff = end_time - start_time
                        time_diff_cell = row[6]
                        time_diff_cell.value = str(time_diff)
                        break

                workbook.save(self.excel_file)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al registrar la hora de reparación en Excel: {e}")

    def log_incidence_to_excel(self, block_name, date_str, time_str, incidence_text):
        if self.excel_file and os.path.exists(self.excel_file):
            try:
                workbook = load_workbook(self.excel_file)
                sheet = workbook.active

                headers = [cell.value for cell in sheet[1]]
                expected_headers = ["Bloque", "Incidencia", "Fecha", "Hora", "Turno", "Hora de Reparación", "Tiempo de Reparación", "MTBF"]
                if headers != expected_headers:
                    sheet.delete_rows(1, 1)
                    sheet.insert_rows(1)
                    for idx, header in enumerate(expected_headers):
                        sheet.cell(row=1, column=idx + 1, value=header)

                time_obj = datetime.strptime(time_str, "%H:%M:%S").time()
                if time(6, 0) <= time_obj < time(18, 0):
                    turno = "Mañana"
                else:
                    turno = "Noche"

                mtbf_value = self.calculate_mtbf(block_name)

                new_row = [block_name, incidence_text, date_str, time_str, turno, "", "", mtbf_value]
                sheet.append(new_row)

                workbook.save(self.excel_file)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al registrar la incidencia en Excel: {e}")
