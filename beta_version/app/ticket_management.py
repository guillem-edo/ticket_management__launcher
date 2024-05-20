# app/ticket_management.py
import json
import os
from datetime import datetime
from collections import Counter, defaultdict
from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import QMainWindow, QVBoxLayout, QWidget, QHBoxLayout, QPushButton, QLineEdit, QFileDialog, QTableWidget, QTableWidgetItem, QListWidget, QLabel, QTabWidget, QStatusBar, QInputDialog, QMessageBox, QAbstractItemView, QListWidgetItem, QApplication
from PyQt5.QtCore import QTimer, Qt, QRect
from PyQt5.QtGui import QFont, QColor
from app.dialogs import AdvancedFilterDialog, TopIncidentsDialog, GraphDialog
from app.admin_dialog import AdminDialog

class TicketManagement(QMainWindow):
    def __init__(self, user):
        super().__init__()
        self.user = user
        self.excel_file = None
        self.incidencias = {
            "WC47 NACP": ["Etiquetadora", "Fallo en elevador", "No atornilla tapa", "Fallo tolva",
                        "Fallo en paletizador", "No coge placa", "Palet atascado en la curva",
                        "Ascensor no sube", "No pone tornillo", "Fallo tornillo", "AOI no detecta pieza",
                        "No atornilla clips", "Fallo fijador tapa", "Secuencia atornillador",
                        "Fallo atornillador", "Fallo cámara visión"],
            "WC48 P5F": ["Etiquetadora", "AOI (fallo etiqueta)", "AOI (malla)", "Cámara no detecta Pcb", "Cámara no detecta skeleton",
                        "Cámara no detecta foams", "Cámara no detecta busbar", "Cámara no detecta foam derecho", "No detecta presencia power CP",
                        "Tornillo atascado en tolva", "Cámara no detecta Power CP", "Cámara no detecta Top cover", "Detección de sealling mal puesto",
                        "Robot no coge busbar", "Fallo etiqueta", "Power atascado en prensa, cuesta sacar", "No coloca bien el sealling"],
            "WC49 P5H": ["La cámara no detecta Busbar", "La cámara no detecta Top Cover", "Screw K30 no lo detecta puesto", "Atasco tuerca",
                        "Tornillo atascado", "Etiquetadora", "Detección de sealling mal puesto", "No coloca bien el sealling", "Power atascado en prensa, cuesta sacar",
                        "No lee QR"],
            "WV50 FILTER": ["Fallo cámara ferrite", "NOK Soldadura Plástico", "NOK Soldadura metal", "Traza", "NOK Soldad. Plástico+Metal", "Robot no coloca bien filter en palet",
                            "No coloca bien la pcb", "QR desplazado", "Core enganchado", "Robot no coge PCB", "Fallo atornillador", "Pieza enganchada en HV Test", "Cover atascado",
                            "Robot no coloca bien ferrita", "No coloca bien el core", "Fallo Funcional", "Fallo visión core", "Fallo cámara cover", "Repeat funcional", "Fallo cámara QR",
                            "No coloca bien foam"],
            "SPL": ["Sensor de PCB detecta que hay placa cuando no la hay", "No detecta marcas Power", "Colisión placas", "Fallo dispensación glue", "Marco atascado en parte inferior",
                    "Soldadura defectuosa", "Error en sensor de salida"]
        }
        self.blocks = user.blocks if not user.is_admin else list(self.incidencias.keys())
        self.last_incidence_labels = {}
        self.incidences_count = {block: 0 for block in self.incidencias.keys()}
        self.pending_incidents = []
        self.filtered_incidents_data = {}
        self.incident_details = {}
        self.config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.txt")
        self.state_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "incidence_state.json")
        self.initUI()
        self.load_last_excel_file()
        self.load_incidence_state()

    def initUI(self):
        self.setWindowTitle(f"Ticket Management - {self.user.username}")
        self.resize(1200, 800)
        self.center_window_app()

        main_layout = QHBoxLayout()
        central_widget = QWidget()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

        self.tabWidget = QTabWidget(self)
        main_layout.addWidget(self.tabWidget, 2)

        for name in self.blocks:
            self.create_tab(name, self.incidencias[name])

        right_layout = QVBoxLayout()
        select_excel_button = QPushButton("Seleccionar Archivo Excel", self)
        select_excel_button.clicked.connect(self.select_excel_file)
        right_layout.addWidget(select_excel_button)

        self.excel_path_display = QLineEdit(self)
        self.excel_path_display.setReadOnly(True)
        right_layout.addWidget(self.excel_path_display)

        self.toggle_excel_view_button = QPushButton("Ver Excel Completo", self)
        self.toggle_excel_view_button.clicked.connect(self.toggle_excel_view)
        right_layout.addWidget(self.toggle_excel_view_button)
        self.excel_view_mode = "completo"

        self.table_widget = QTableWidget(self)
        right_layout.addWidget(self.table_widget)

        self.global_incidence_list = QListWidget(self)
        self.global_incidence_list.setFixedSize(600, 300)
        self.global_incidence_list.setStyleSheet("QListWidget { background-color: #f0f0f0; border: 1px solid #ccc; }")
        right_layout.addWidget(self.global_incidence_list)

        filter_button = QPushButton("Filtro Avanzado", self)
        filter_button.clicked.connect(self.open_advanced_filter_dialog)
        right_layout.addWidget(filter_button)

        self.top_incidents_label = QLabel("Incidencias Más Relevantes")
        right_layout.addWidget(self.top_incidents_label)

        self.view_details_button = QPushButton("Ver Detalles")
        self.view_details_button.clicked.connect(self.open_top_incidents_dialog)
        right_layout.addWidget(self.view_details_button)

        self.graph_button = QPushButton("Ver Gráficos")
        self.graph_button.clicked.connect(self.open_graph_dialog)
        right_layout.addWidget(self.graph_button)

        if self.user.is_admin:
            admin_button = QPushButton("Administrar Incidencias")
            admin_button.setStyleSheet("background-color: #f0ad4e; color: white; padding: 10px 20px; font-size: 16px;")
            admin_button.clicked.connect(self.open_admin_dialog)
            right_layout.addWidget(admin_button)

        right_layout.addStretch()

        right_widget = QWidget()
        right_widget.setLayout(right_layout)
        main_layout.addWidget(right_widget, 1)

        self.status_bar = QStatusBar(self)
        self.setStatusBar(self.status_bar)

        self.detailed_messages_tab = QWidget()
        self.detailed_messages_layout = QVBoxLayout(self.detailed_messages_tab)
        self.tabWidget.addTab(self.detailed_messages_tab, "Mensajes detallados")
        self.detailed_messages_list = QListWidget()
        self.detailed_messages_list.setStyleSheet("QListWidget { background-color: #f0f0f0; border: 1px solid #ccc; }")
        self.detailed_messages_layout.addWidget(self.detailed_messages_list)

        timer = QTimer(self)
        timer.timeout.connect(self.update_status_bar)
        timer.start(1000)

        self.apply_styles()

    def center_window_app(self):
        screen_rect_app = QApplication.desktop().availableGeometry()
        window_width = self.width()
        window_height = self.height()
        x = (screen_rect_app.width() - window_width) // 2
        y = (screen_rect_app.height() - window_height) // 2
        self.setGeometry(QRect(x, y, window_width, window_height))

    def open_advanced_filter_dialog(self):
        self.filter_dialog = AdvancedFilterDialog(self)
        self.filter_dialog.exec_()

    def open_top_incidents_dialog(self):
        self.top_incidents_dialog = TopIncidentsDialog(self, incident_details=self.incident_details)
        self.top_incidents_dialog.exec_()

    def open_graph_dialog(self):
        self.graph_dialog = GraphDialog(self, data=self.filtered_incidents_data)
        self.graph_dialog.exec_()

    def open_admin_dialog(self):
        self.admin_dialog = AdminDialog(self, incidencias=self.incidencias)
        self.admin_dialog.exec_()

    def toggle_excel_view(self):
        if self.excel_view_mode == "completo":
            self.excel_view_mode = "filtrado"
            self.toggle_excel_view_button.setText("Ver Excel Completo")
        else:
            self.excel_view_mode = "completo"
            self.toggle_excel_view_button.setText("Ver Excel Filtrado")
        self.update_excel_table()

    def get_filtered_incidents(self, start_dt, end_dt, selected_block):
        if not self.excel_file or not os.path.exists(self.excel_file):
            return {}, {}

        workbook = load_workbook(self.excel_file)
        sheet = workbook.active

        filtered_counts = {block: {'count': 0, 'incidences': []} for block in self.incidencias.keys()}
        trends = {block: defaultdict(int) for block in self.incidencias.keys()}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            date_str, time_str, *incidences = row
            if date_str and time_str:
                try:
                    row_datetime = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
                    if start_dt <= row_datetime <= end_dt:
                        hour_str = row_datetime.strftime("%Y-%m-%d %H:00:00")
                        for i, incidence in enumerate(incidences):
                            if incidence != "-":
                                block = list(self.incidencias.keys())[i]
                                if selected_block == "Todos" or selected_block == block:
                                    filtered_counts[block]['count'] += 1
                                    filtered_counts[block]['incidences'].append(incidence)
                                    trends[block][hour_str] += 1
                except ValueError:
                    continue

        for block, data in filtered_counts.items():
            if data['incidences']:
                most_common_incidence, _ = Counter(data['incidences']).most_common(1)[0]
                data['most_common_incidence'] = most_common_incidence
            else:
                data['most_common_incidence'] = "N/A"

        self.filtered_incidents_data = filtered_counts

        if selected_block != "Todos":
            filtered_counts = {selected_block: filtered_counts[selected_block]}
            trends = {selected_block: trends[selected_block]}

        return filtered_counts, trends

    def create_tab(self, name, incidences):
        tab = QWidget()
        self.tabWidget.addTab(tab, name)
        layout = QVBoxLayout(tab)

        title = QLabel(name)
        title.setFont(QFont("Arial", 14, QFont.Bold))
        layout.addWidget(title)

        last_incidence_label = QLabel("Última incidencia: Ninguna")
        last_incidence_label.setFont(QFont("Arial", 12))
        layout.addWidget(last_incidence_label)
        self.last_incidence_labels[name] = last_incidence_label

        list_widget = QListWidget()
        font = QFont("Arial", 12)
        list_widget.setFont(font)
        list_widget.setStyleSheet("""
            QListWidget {
                background-color: #f9f9f9;
                border: 1px solid #ddd;
                padding: 10px;
            }
            QListWidget::item {
                padding: 10px;
                border-bottom: 1px solid #ddd;
            }
            QListWidget::item:selected {
                background-color: #4CAF50;
                color: white;
            }
        """)
        for incidence in incidences:
            item = QListWidgetItem(incidence)
            list_widget.addItem(item)

        layout.addWidget(list_widget)

        confirm_button = QPushButton("Confirmar Incidencia")
        confirm_button.setFont(QFont("Arial", 12))
        confirm_button.setFixedSize(500, 40)
        confirm_button.setStyleSheet("background-color: #4CAF50; color: white;")
        confirm_button.clicked.connect(lambda: self.confirm_incidence(name, list_widget))
        layout.addWidget(confirm_button)

    def confirm_incidence(self, block_name, list_widget):
        current_item = list_widget.currentItem()
        if current_item:
            incidence_text = current_item.text()
            timestamp = datetime.now()
            date_str = timestamp.strftime("%Y-%m-%d")
            time_str = timestamp.strftime("%H:%M:%S")
            self.last_incidence_labels[block_name].setText(f"Incidencia confirmada: {incidence_text} a las {time_str}")
            self.log_incidence_to_excel(block_name, date_str, time_str, incidence_text)

            QMessageBox.information(self, "Confirmación", "Incidencia confirmada.")

            fixing_label = QLabel("Fixing")
            fixing_label.setStyleSheet("color: red; font-weight: bold; font-size: 14px;")
            correct_button = QPushButton("Correct")
            correct_button.setStyleSheet("background-color: green; color: white; font-weight: bold; font-size: 14px;")
            correct_button.setFixedSize(100, 30)
            details_button = QPushButton("Añadir Detalles")
            details_button.setStyleSheet("background-color: blue; color: white; font-weight: bold; font-size: 14px;")
            details_button.setFixedSize(150, 30)

            item_widget = QWidget()
            item_layout = QVBoxLayout(item_widget)
            label_layout = QHBoxLayout()
            label_layout.addWidget(QLabel(f"{block_name}: {incidence_text} a las {time_str} del {date_str}"))
            item_layout.addLayout(label_layout)

            buttons_layout = QHBoxLayout()
            buttons_layout.addStretch()
            buttons_layout.addWidget(fixing_label)
            buttons_layout.addWidget(correct_button)
            buttons_layout.addWidget(details_button)
            item_layout.addLayout(buttons_layout)

            list_item = QListWidgetItem()
            list_item.setSizeHint(item_widget.sizeHint())

            self.global_incidence_list.addItem(list_item)
            self.global_incidence_list.setItemWidget(list_item, item_widget)

            correct_button.clicked.connect(lambda: self.mark_incidence_as_fixed(block_name, incidence_text, fixing_label, correct_button, details_button, date_str, time_str))
            details_button.clicked.connect(lambda: self.add_incidence_details(block_name, incidence_text, date_str, time_str))

            QTimer.singleShot(60000, lambda: self.remind_user_to_fix(block_name, incidence_text, fixing_label, correct_button, details_button, date_str, time_str))

            self.update_excel_table()
            self.update_top_incidents()
        else:
            QMessageBox.warning(self, "Ninguna Incidencia Seleccionada", "Selecciona una incidencia para confirmar.")

    def remind_user_to_fix(self, block_name, incidence_text, fixing_label, correct_button, details_button, date_str, time_str):
        if fixing_label.text() == "Fixing":
            msg_box = QMessageBox(self)
            msg_box.setIcon(QMessageBox.Warning)
            msg_box.setText(f'La incidencia "{incidence_text}" sigue en estado "Fixing".')
            msg_box.setInformativeText("¿Deseas marcarla como 'Pendiente' o continuar?")
            msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            yes_button = msg_box.button(QMessageBox.Yes)
            yes_button.setText("Pendiente")
            no_button = msg_box.button(QMessageBox.No)
            no_button.setText("Continuar")
            msg_box.exec_()

            if msg_box.clickedButton() == yes_button:
                fixing_label.setText("Pendiente")
                fixing_label.setStyleSheet("color: orange; font-weight: bold; font-size: 14px;")
                self.pending_incidents.append((block_name, incidence_text, date_str, time_str))
                self.redirect_to_pending_incidence(block_name, incidence_text, fixing_label, correct_button, details_button, date_str, time_str)
            else:
                QTimer.singleShot(60000, lambda: self.remind_user_to_fix(block_name, incidence_text, fixing_label, correct_button, details_button, date_str, time_str))

    def redirect_to_pending_incidence(self, block_name, incidence_text, fixing_label, correct_button, details_button, date_str, time_str):
        for i in range(self.global_incidence_list.count()):
            item = self.global_incidence_list.item(i)
            item_widget = self.global_incidence_list.itemWidget(item)
            if item_widget:
                label = item_widget.findChild(QLabel)
                if label and incidence_text in label.text():
                    self.global_incidence_list.setCurrentRow(i)
                    self.global_incidence_list.scrollToItem(item, QAbstractItemView.PositionAtCenter)
                    break

    def mark_incidence_as_fixed(self, block_name, incidence_text, fixing_label, correct_button, details_button, date_str, time_str):
        fixing_label.setText("Reparada")
        fixing_label.setStyleSheet("color: green; font-weight: bold; font-size: 14px;")
        correct_button.setEnabled(False)

        repair_time_str = datetime.now().strftime("%H:%M:%S")
        self.log_repair_time_to_excel(block_name, date_str, time_str, repair_time_str)
        self.update_excel_table()

    def add_incidence_details(self, block_name, incidence_text, date_str, time_str):
        detail_text, ok = QInputDialog.getMultiLineText(self, "Añadir Detalles", "Escribe los detalles de la incidencia:")
        if ok and detail_text:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            detail_message = f"{timestamp} - {block_name}: {incidence_text} ({date_str} {time_str})\nDetalles: {detail_text}"

            self.detailed_messages_list.addItem(detail_message)

    def log_repair_time_to_excel(self, block_name, date_str, time_str, repair_time_str):
        if self.excel_file and os.path.exists(self.excel_file):
            try:
                workbook = load_workbook(self.excel_file)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2):
                    if row[0].value == date_str and row[1].value == time_str:
                        repair_time_cell = sheet.cell(row=row[0].row, column=len(self.incidencias) + 3)
                        repair_time_cell.value = repair_time_str

                        start_time = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
                        end_time = datetime.strptime(f"{date_str} {repair_time_str}", "%Y-%m-%d %H:%M:%S")
                        time_diff = end_time - start_time
                        time_diff_cell = sheet.cell(row=row[0].row, column=len(self.incidencias) + 4)
                        time_diff_cell.value = str(time_diff)
                        break

                workbook.save(self.excel_file)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al registrar la hora de reparación en Excel: {e}")

    def select_excel_file(self):
        file_dialog = QFileDialog()
        file_dialog.setNameFilter("Archivos Excel (*.xlsx)")
        if file_dialog.exec_():
            self.excel_file = file_dialog.selectedFiles()[0]
            self.excel_path_display.setText(self.excel_file)
            with open(self.config_file, "w") as config:
                config.write(self.excel_file)
            self.update_excel_table()
            self.update_top_incidents()

    def create_excel_if_not_exists(self, file_path):
        if not os.path.exists(file_path):
            workbook = Workbook()
            sheet = workbook.active
            headers = ["Fecha", "Hora"] + list(self.incidencias.keys()) + ["Hora de Reparación", "Tiempo de Reparación"]
            sheet.append(headers)
            workbook.save(file_path)

    def log_incidence_to_excel(self, block_name, date_str, time_str, incidence_text):
        if self.excel_file and os.path.exists(self.excel_file):
            try:
                workbook = load_workbook(self.excel_file)
                sheet = workbook.active

                headers = [cell.value for cell in sheet[1]]
                expected_headers = ["Fecha", "Hora"] + list(self.incidencias.keys()) + ["Hora de Reparación", "Tiempo de Reparación"]
                if headers != expected_headers:
                    sheet.delete_rows(1, 1)
                    sheet.insert_rows(1)
                    for idx, header in enumerate(expected_headers):
                        sheet.cell(row=1, column=idx + 1, value=header)

                new_row = [date_str, time_str] + ["-"] * len(self.incidencias) + ["", ""]
                block_index = list(self.incidencias.keys()).index(block_name) + 2
                new_row[block_index] = incidence_text
                sheet.append(new_row)

                workbook.save(self.excel_file)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al registrar la incidencia en Excel: {e}")

    def update_status_bar(self):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.statusBar().showMessage(f"Fecha y Hora Actual: {current_time}")

    def load_last_excel_file(self):
        if os.path.exists(self.config_file):
            with open(self.config_file, "r") as config:
                file_path = config.read().strip()
                if os.path.exists(file_path):
                    self.excel_file = file_path
                    self.excel_path_display.setText(file_path)
                    self.update_excel_table()
                    self.update_top_incidents()

    def update_excel_table(self):
        if self.excel_file and os.path.exists(self.excel_file):
            try:
                workbook = load_workbook(self.excel_file)
                sheet = workbook.active
                rows_to_display = []

                if self.excel_view_mode == "completo":
                    rows_to_display = list(sheet.iter_rows(min_row=2, values_only=True))
                else:
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row[list(self.incidencias.keys()).index(self.blocks[0]) + 2] != "-":
                            rows_to_display.append(row)

                self.table_widget.setRowCount(len(rows_to_display))
                self.table_widget.setColumnCount(sheet.max_column)

                headers = [cell.value for cell in sheet[1]]
                self.table_widget.setHorizontalHeaderLabels(headers)

                for row_idx, row in enumerate(rows_to_display):
                    for col_idx, cell_value in enumerate(row):
                        item = QTableWidgetItem(str(cell_value) if cell_value is not None else "")
                        self.table_widget.setItem(row_idx, col_idx, item)

                self.style_excel_table()

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al actualizar la tabla: {e}")

    def style_excel_table(self):
        for row in range(self.table_widget.rowCount()):
            for col in range(self.table_widget.columnCount()):
                item = self.table_widget.item(row, col)
                if item:
                    item.setFont(QFont("Arial", 10))
                    if row % 2 == 0:
                        item.setBackground(QColor(240, 240, 240))
                    else:
                        item.setBackground(QColor(255, 255, 255))
                    item.setTextAlignment(Qt.AlignCenter)

    def update_top_incidents(self):
        self.incidences_count = {block: 0 for block in self.incidencias.keys()}
        self.incident_details = {block: Counter() for block in self.incidencias.keys()}

        if self.excel_file and os.path.exists(self.excel_file):
            workbook = load_workbook(self.excel_file)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) == len(self.incidencias) + 4:
                    for i, block in enumerate(list(self.incidencias.keys()), start=2):
                        if row[i] and row[i] != "-":
                            self.incidences_count[block] += 1
                            self.incident_details[block][row[i]] += 1

    def apply_styles(self):
        title_font = QFont("Arial", 14, QFont.Bold)
        normal_font = QFont("Arial", 12)

        for i in range(self.tabWidget.count()):
            tab = self.tabWidget.widget(i)
            for widget in tab.findChildren(QLabel):
                if "Última incidencia" in widget.text():
                    widget.setFont(normal_font)
                else:
                    widget.setFont(title_font)

    def save_incidence_state(self):
        incidences = []
        for i in range(self.global_incidence_list.count()):
            item = self.global_incidence_list.item(i)
            item_widget = self.global_incidence_list.itemWidget(item)
            if item_widget:
                labels = item_widget.findChildren(QLabel)
                if labels:
                    incidence_text = labels[0].text()
                    status = labels[1].text() if len(labels) > 1 else "Fixing"
                    incidences.append({
                        "text": incidence_text,
                        "status": status
                    })
        state = {
            "user": self.user.username,
            "incidences": incidences
        }
        with open(self.state_file, "w") as file:
            json.dump(state, file, indent=4)

    def load_incidence_state(self):
        if os.path.exists(self.state_file):
            with open(self.state_file, "r") as file:
                state = json.load(file)
                if state["user"] == self.user.username:
                    for incidence in state["incidences"]:
                        incidence_text = incidence["text"]
                        status = incidence["status"]

                        # Extraer date_str y time_str del texto de la incidencia
                        text_parts = incidence_text.split(" ")
                        date_str = text_parts[-3]
                        time_str = text_parts[-2]

                        fixing_label = QLabel(status)
                        if status == "Fixing":
                            fixing_label.setStyleSheet("color: red; font-weight: bold; font-size: 14px;")
                        elif status == "Pendiente":
                            fixing_label.setStyleSheet("color: orange; font-weight: bold; font-size: 14px;")
                        elif status == "Reparada":
                            fixing_label.setStyleSheet("color: green; font-weight: bold; font-size: 14px;")
                        correct_button = QPushButton("Correct")
                        correct_button.setStyleSheet("background-color: green; color: white; font-weight: bold; font-size: 14px;")
                        correct_button.setFixedSize(100, 30)
                        details_button = QPushButton("Añadir Detalles")
                        details_button.setStyleSheet("background-color: blue; color: white; font-weight: bold; font-size: 14px;")
                        details_button.setFixedSize(150, 30)
                        item_widget = QWidget()
                        item_layout = QVBoxLayout(item_widget)
                        label_layout = QHBoxLayout()
                        label_layout.addWidget(QLabel(incidence_text))
                        item_layout.addLayout(label_layout)
                        buttons_layout = QHBoxLayout()
                        buttons_layout.addStretch()
                        buttons_layout.addWidget(fixing_label)
                        buttons_layout.addWidget(correct_button)
                        buttons_layout.addWidget(details_button)
                        item_layout.addLayout(buttons_layout)
                        list_item = QListWidgetItem()
                        list_item.setSizeHint(item_widget.sizeHint())
                        self.global_incidence_list.addItem(list_item)
                        self.global_incidence_list.setItemWidget(list_item, item_widget)
                        correct_button.clicked.connect(lambda: self.mark_incidence_as_fixed(self.user.blocks[0], incidence_text, fixing_label, correct_button, details_button, date_str, time_str))
                        details_button.clicked.connect(lambda: self.add_incidence_details(self.user.blocks[0], incidence_text, date_str, time_str))
                        if status == "Fixing" or status == "Pendiente":
                            QTimer.singleShot(60000, lambda: self.remind_user_to_fix(self.user.blocks[0], incidence_text, fixing_label, correct_button, details_button, date_str, time_str))
