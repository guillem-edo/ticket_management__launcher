import sys
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QWidget, QPushButton, QMessageBox, QTabWidget, QLabel,
    QListWidget, QListWidgetItem, QStatusBar, QLineEdit, QFileDialog, QHBoxLayout, QTableWidget, QTableWidgetItem
)
from PyQt5.QtCore import QTimer, Qt, pyqtSignal, QRect
from PyQt5.QtGui import QFont

from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

class LoginWindow(QWidget):
    login_successful = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Inicio de Sesión")
        self.resize(280, 360)
        self.center_window()

        layout = QVBoxLayout()
        logo_label = QLabel("Inicio de Sesión")
        layout.addWidget(logo_label)

        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("Usuario")
        layout.addWidget(self.username_input)

        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Contraseña")
        self.password_input.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.password_input)

        login_button = QPushButton("Confirmar")
        login_button.clicked.connect(self.login)
        layout.addWidget(login_button)

        self.setLayout(layout)

    def center_window(self):
        screen_rect = QApplication.desktop().availableGeometry()
        x = (screen_rect.width() - self.width()) // 2
        y = (screen_rect.height() - self.height()) // 2
        self.setGeometry(QRect(x, y, self.width(), self.height()))

    def login(self):
        username = self.username_input.text()
        password = self.password_input.text()
        if username == "pideu" and password == "1111":
            self.login_successful.emit()
            self.close()
        else:
            QMessageBox.warning(self, "Error de Inicio de Sesión", "Usuario o contraseña incorrectos")

class TicketManagement(QMainWindow):
    config_file = "config.txt"

    def __init__(self):
        super().__init__()
        self.excel_file = None
        self.blocks = ["WC47 NACP", "WC48 P5F", "WC49 P5H", "WV50 FILTER", "SPL"]
        self.last_incidence_labels = {}
        self.incidencias = {
            "WC47 NACP": ["Etiquetadora", "Fallo en elevador", "No atornilla tapa", "Fallo tolva",
                        "Fallo en paletizador", "No coge placa", "Palet atascado en la curva",
                        "Ascensor no sube", "No pone tornillo", "Fallo tornillo", "AOI no detecta pieza",
                        "No atornilla clips", "Fallo fijador tapa", "Secuencia atornillador",
                        "Fallo atornillador", "Fallo cámara visión"],
            "WC48 P5F": ["Etiquetadora", "AOI (fallo etiqueta)", "AOI (malla)", "Cámara no detecta Pcb", "Cámara no detecta skeleton",
                        "Cámara no detecta foams", "Cámara no detecta busbar", "Cámara no detecta foam derecho", "No detecta presencia power CP",
                        "Tornillo atascado en tolva", "Cámara no detecta Power CP", "Cámara no detecta Top cover", "Detección de sealling mal puesto",
                        "Robot no coge busbar", "Fallo etiqueta", "Power atascado en prensa, cuesta sacar", "No coloca bien el sealling"],
            "WC49 P5H": ["La cámara no detecta Busbar", "La cámara no detecta Top Cover", "Screw K30 no lo detecta puesto", "Atasco tuerca",
                        "Tornillo atascado", "Etiquetadora", "Detección de sealling mal puesto", "No coloca bien el sealling", "Power atascado en prensa, cuesta sacar",
                        "No lee QR"],
            "WV50 FILTER": ["Fallo cámara ferrite", "NOK Soldadura Plástico", "NOK Soldadura metal", "Traza", "NOK Soldad. Plástico+Metal", "Robot no coloca bien filter en palet",
                            "No coloca bien la pcb", "QR desplazado", "Core enganchado", "Robot no coge PCB", "Fallo atornillador", "Pieza enganchada en HV Test", "Cover atascado",
                            "Robot no coloca bien ferrita", "No coloca bien el core", "Fallo Funcional", "Fallo visión core", "Fallo cámara cover", "Repeat funcional", "Fallo cámara QR",
                            "No coloca bien foam"],
            "SPL": ["Sensor de PCB detecta que hay placa cuando no la hay", "No detecta marcas Power", "Colisión placas", "Fallo dispensación glue", "Marco atascado en parte inferior",
                    "Soldadura defectuosa", "Error en sensor de salida"]
        }
        self.incidences_count = {block: 0 for block in self.blocks}
        self.initUI()
        self.load_last_excel_file()

    def initUI(self):
        self.setWindowTitle("Ticket Management")
        self.setGeometry(200, 200, 1200, 800)  # Tamaño de ventana ajustado

        main_layout = QHBoxLayout()
        central_widget = QWidget()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

        # Sección de Tabs de Incidencias
        self.tabWidget = QTabWidget(self)
        main_layout.addWidget(self.tabWidget, 2)

        for name, incidences in self.incidencias.items():
            self.create_tab(name, incidences)

        # Sección para Seleccionar Excel y mostrar incidencias confirmadas
        right_layout = QVBoxLayout()
        select_excel_button = QPushButton("Seleccionar Archivo Excel", self)
        select_excel_button.clicked.connect(self.select_excel_file)
        right_layout.addWidget(select_excel_button)

        self.excel_path_display = QLineEdit(self)
        self.excel_path_display.setReadOnly(True)
        right_layout.addWidget(self.excel_path_display)

        self.table_widget = QTableWidget(self)
        self.table_widget.setFixedSize(600, 300)  # Tamaño fijo para la tabla
        right_layout.addWidget(self.table_widget)

        self.global_incidence_list = QListWidget(self)
        right_layout.addWidget(self.global_incidence_list)

        # Sección del gráfico
        graph_layout = QVBoxLayout()
        self.figure = Figure()
        self.canvas = FigureCanvas(self.figure)
        graph_layout.addWidget(self.canvas)

        right_layout.addLayout(graph_layout)  # Añadir el layout del gráfico al layout de la derecha

        right_layout.addStretch()

        right_widget = QWidget()
        right_widget.setLayout(right_layout)
        main_layout.addWidget(right_widget, 1)

        # Barra de estado para mostrar mensajes
        self.status_bar = QStatusBar(self)
        self.setStatusBar(self.status_bar)

        timer = QTimer(self)
        timer.timeout.connect(self.update_status_bar)
        timer.start(1000)

        self.apply_styles()

    def create_tab(self, name, incidences):
        tab = QWidget()
        self.tabWidget.addTab(tab, name)
        layout = QVBoxLayout(tab)

        title = QLabel(f"Incidencias - {name}")
        layout.addWidget(title)

        last_incidence_label = QLabel("Última incidencia: Ninguna")
        layout.addWidget(last_incidence_label)
        self.last_incidence_labels[name] = last_incidence_label

        list_widget = QListWidget()
        for incidence in incidences:
            item = QListWidgetItem(incidence)
            list_widget.addItem(item)

        layout.addWidget(list_widget)

        confirm_button = QPushButton("Confirmar Incidencia")
        confirm_button.clicked.connect(lambda: self.confirm_incidence(name, list_widget))
        layout.addWidget(confirm_button)

    def confirm_incidence(self, block_name, list_widget):
        current_item = list_widget.currentItem()
        if current_item:
            incidence_text = current_item.text()
            timestamp = datetime.now()
            date_str = timestamp.strftime("%Y-%m-%d")
            time_str = timestamp.strftime("%H:%M:%S")
            self.last_incidence_labels[block_name].setText(f"Incidencia confirmada: {incidence_text} a las {time_str}")
            self.log_incidence_to_excel(block_name, date_str, time_str, incidence_text)

            # Mostrar el cuadro de diálogo de confirmación
            QMessageBox.information(self, "Confirmación", "Incidencia confirmada.")

            # Actualizar la tabla y la información de la incidencia confirmada
            self.update_excel_table()

            # Actualizar el conteo de incidencias y la gráfica
            self.update_graph()

            # Añadir incidencia a la lista global
            self.global_incidence_list.addItem(f"{block_name}: {incidence_text} a las {time_str} del {date_str}")
        else:
            QMessageBox.warning(self, "Ninguna Incidencia Seleccionada", "Selecciona una incidencia para confirmar.")

    def select_excel_file(self):
        file_dialog = QFileDialog()
        file_dialog.setNameFilter("Archivos Excel (*.xlsx)")
        if file_dialog.exec_():
            file_path = file_dialog.selectedFiles()[0]
            self.excel_file = file_path
            self.excel_path_display.setText(file_path)
            with open(self.config_file, "w") as config:
                config.write(file_path)
            self.update_excel_table()
            self.update_graph()

    def create_excel_if_not_exists(self, file_path):
        if not os.path.exists(file_path):
            workbook = Workbook()
            sheet = workbook.active
            headers = ["Fecha", "Hora"] + self.blocks  # Cabeceras con los nombres de los bloques
            sheet.append(headers)
            workbook.save(file_path)

    def log_incidence_to_excel(self, block_name, date_str, time_str, incidence_text):
        if self.excel_file and os.path.exists(self.excel_file):
            try:
                workbook = load_workbook(self.excel_file)
                sheet = workbook.active

                # Verificar si el archivo ya tiene las cabeceras correctas
                headers = [cell.value for cell in sheet[1]]
                if headers != ["Fecha", "Hora"] + self.blocks:
                    sheet.delete_rows(1, 1)  # Eliminar la primera fila incorrecta
                    sheet.insert_rows(1)  # Insertar una nueva primera fila
                    for idx, header in enumerate(["Fecha", "Hora"] + self.blocks):
                        sheet.cell(row=1, column=idx + 1, value=header)

                new_row = [date_str, time_str] + ["-"] * len(self.blocks)
                block_index = self.blocks.index(block_name) + 2
                new_row[block_index] = incidence_text
                sheet.append(new_row)

                workbook.save(self.excel_file)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al registrar la incidencia en Excel: {e}")

    def update_status_bar(self):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.statusBar().showMessage(f"Fecha y Hora Actual: {current_time}")

    def load_last_excel_file(self):
        if os.path.exists(self.config_file):
            with open(self.config_file, "r") as config:
                file_path = config.read().strip()
                if os.path.exists(file_path):
                    self.excel_file = file_path
                    self.excel_path_display.setText(file_path)
                    self.update_excel_table()
                    self.update_graph()

    def update_excel_table(self):
        if self.excel_file and os.path.exists(self.excel_file):
            try:
                workbook = load_workbook(self.excel_file)
                sheet = workbook.active
                self.table_widget.setRowCount(sheet.max_row)
                self.table_widget.setColumnCount(sheet.max_column)

                headers = [cell.value for cell in sheet[1]]  # Leer las cabeceras
                self.table_widget.setHorizontalHeaderLabels(headers)  # Configurar las cabeceras

                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):  # Empezar desde la segunda fila
                    for col_idx, cell_value in enumerate(row):
                        item = QTableWidgetItem(str(cell_value) if cell_value is not None else "")
                        self.table_widget.setItem(row_idx - 1, col_idx, item)  # Ajustar el índice de fila

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al actualizar la tabla: {e}")

    def update_graph(self):
        # Reiniciar el conteo de incidencias
        self.incidences_count = {block: 0 for block in self.blocks}
        
        if self.excel_file and os.path.exists(self.excel_file):
            workbook = load_workbook(self.excel_file)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Ignorar la fila de cabeceras
                for i, block in enumerate(self.blocks, start=2):
                    if row[i] != "-":  # Si hay una incidencia registrada
                        self.incidences_count[block] += 1

        self.figure.clear()
        ax = self.figure.add_subplot(111)
        ax.bar(self.incidences_count.keys(), self.incidences_count.values())
        ax.set_xlabel('Bloques')
        ax.set_ylabel('Incidencias Confirmadas')
        ax.set_title('Conteo de Incidencias Confirmadas por Bloque')

        # Reducir nombres del eje X
        short_labels = [block.split()[0] for block in self.blocks]
        ax.set_xticklabels(short_labels, rotation=45, ha="right")

        self.canvas.draw()

    def apply_styles(self):
        title_font = QFont("Arial", 16, QFont.Bold)
        normal_font = QFont("Arial", 12)

        for i in range(self.tabWidget.count()):
            tab = self.tabWidget.widget(i)
            for widget in tab.findChildren(QLabel):
                if "Última incidencia" in widget.text():
                    widget.setFont(normal_font)
                else:
                    widget.setFont(title_font)

if __name__ == "__main__":
    app = QApplication(sys.argv)

    login_window = LoginWindow()
    ticket_management = TicketManagement()

    def on_login_successful():
        ticket_management.show()

    login_window.login_successful.connect(on_login_successful)
    login_window.show()

    sys.exit(app.exec_())