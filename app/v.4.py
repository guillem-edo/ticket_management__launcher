import sys
import json
import csv
import os

from openpyxl import Workbook, load_workbook

from datetime import datetime
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QWidget, QPushButton, QMessageBox, QTabWidget, QLabel, QListWidget, QListWidgetItem, QStatusBar, QHBoxLayout, QLineEdit
)
from PyQt5.QtGui import QFont, QIcon, QPixmap
from PyQt5.QtNetwork import QNetworkAccessManager, QNetworkRequest, QNetworkReply
from PyQt5.QtCore import QUrl, QByteArray, QTimer, Qt, pyqtSignal, QRect

class LoginWindow(QWidget):
    login_successful = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Inicio de Sesión")
        self.resize(280,360)

        self.center_window()

        layout = QVBoxLayout()

        # Logo
        logo_label = QLabel()
        pixmap = QPixmap("app/logo.png").scaled(180, 180, Qt.KeepAspectRatio)
        logo_label.setPixmap(pixmap)
        logo_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(logo_label)

        # Etiquetas personalizadas de "Usuario" y "Contraseña"
        user_label = QLabel("Usuario")
        password_label = QLabel("Contraseña")
        
        # Configura estilos con el método setStyleSheet
        self.setStyleSheet("""
            QWidget {
                background-color: #ffffff;
            }
            QLabel {
                font-size: 20px;
                color: #34495e;
                margin-bottom: 5px;
            }
            QLineEdit {
                font-size: 28px;
                padding: 20px;
                margin: 10px 0;
                border: 2px solid #2980b9;
                border-radius: 8px;
                background-color: #f0f4f8;
            }
            QPushButton {
                background-color: #2980b9;
                color: #ffffff;
                border-radius: 8px;
                font-size: 20px;
                height: 60px;
                margin-top: 15px;
            }
        """)

        # Campo de entrada de "Usuario"
        layout.addWidget(user_label)
        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("User")
        layout.addWidget(self.username_input)

        # Campo de entrada de "Contraseña"
        layout.addWidget(password_label)
        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Password")
        self.password_input.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.password_input)

        # Botón de inicio de sesión
        login_button = QPushButton("Confirmar")
        login_button.clicked.connect(self.login)
        layout.addWidget(login_button)

        self.setLayout(layout)

    def center_window(self):
        # Rectángulo del escritorio
        screen_rect = QApplication.desktop().availableGeometry()
        # Coordenadas
        x = (screen_rect.width() - self.width()) // 2
        y = (screen_rect.height() - self.height()) // 2
        # Posicionar la ventana
        self.setGeometry(QRect(x, y, self.width(), self.height()))

    def login(self):
        username = self.username_input.text()
        password = self.password_input.text()
        if username == "ficosa_pideu" and password == "1111":
            self.login_successful.emit()
            self.close()
        else:
            QMessageBox.warning(self, "Error de Inicio de Sesión", "Usuario o contraseña incorrectos")

class TicketManagement(QMainWindow):
    last_incidence_changed = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.last_incidence_labels = {}
        self.initUI()
        self.network_manager = QNetworkAccessManager(self)
        self.network_manager.finished.connect(self.on_network_reply)
        self.last_incidence = None
        self.blocks = ["WC47 NACP", "WC48 POWER 5F", "WC49 POWER 5H", "WV50 FILTER", "SPL"]
        self.excel_file = 'Incidencias_Pideu.xlsx'
        self.create_excel_file()

    def initUI(self):
        # Rectángulo del escritorio
        screen_rect_1 = QApplication.desktop().availableGeometry()
        # Coordenadas
        x = (screen_rect_1.width() - self.width()) // 2
        y = (screen_rect_1.height() - self.height()) // 2
        # Posicionar la ventana
        self.setGeometry(QRect(x, y, self.width(), self.height()))
        self.setWindowTitle("Ticket Management")
        # self.setGeometry(100, 100, 1200, 800)
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f4f6f7;
            }
            QPushButton {
                background-color: #3498db;
                color: white;
                border-radius: 5px;
                padding: 5px 50px;
                font-size: 18px;
                height: 40px;
            }
            QListWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
            QListWidget::item {
                padding: 5px;
                border: 1px solid #dcdcdc;
                border-radius: 3px;
            }
            QTabWidget {
                font-size: 14px;
            }
            QLabel {
                font-size: 14px;
            }
        """)

        self.tabWidget = QTabWidget(self)
        self.tabWidget.setFont(QFont('Arial', 12))
        self.setCentralWidget(self.tabWidget)

        self.status_bar = QStatusBar(self)
        self.setStatusBar(self.status_bar)
        self.update_status_bar()

        timer = QTimer(self)
        timer.timeout.connect(self.update_status_bar)
        timer.start(1000)

        self.incidencias = {
            "WC47 NACP": ["Etiquetadora", "Fallo en elevador", "No atornilla tapa", "Fallo tolva",
                        "Fallo en paletizador", "No coge placa", "Palet atascado en la curva",
                        "Ascensor no sube", "No pone tornillo", "Fallo tornillo", "AOI no detecta pieza",
                        "No atornilla clips", "Fallo fijador tapa", "Secuencia atornillador",
                        "Fallo atornillador", "Fallo cámara visión"],
            "WC48 POWER 5F": ["Etiquetadora","AOI (fallo etiqueta)","AOI (malla)","Cámara no detecta Pcb","Cámara no detecta skeleton",
                        "Cámara no detecta foams","Cámara no detecta busbar","Cámara no detecta foam derecho","No detecta presencia power CP",
                        "Tornillo atascado en tolva","Cámara no detecta Power CP","Cámara no detecta Top cover","Detección de sealling mal puesto",
                        "Robot no coge busbar","Fallo etiqueta","Power atascado en prensa, cuesta sacar","No coloca bien el sealling"],
            "WC49 POWER 5H": ["La cámara no detecta Busbar","La cámara no detecta Top Cover","Screw K30 no lo detecta puesto","Atasco tuerca",
                        "Tornillo atascado","Etiquetadora","Detección de sealling mal puesto","No coloca bien el sealling","Power atascado en prensa, cuesta sacar",
                        "No lee QR"],
            "WV50 FILTER": ["Fallo cámara ferrite","NOK Soldadura Plástico","NOK Soldadura metal","Traza","NOK Soldad. Plástico+Metal","Robot no coloca bien filter en palet",
                        "No coloca bien la pcb","QR desplazado","Core enganchado","Robot no coge PCB","Fallo atornillador","Pieza enganchada en HV Test","Cover atascado",
                        "Robot no coloca bien ferrita","No coloca bien el core","Fallo Funcional","Fallo visión core","Fallo cámara cover","Repeat funcional","Fallo cámara QR",
                        "No coloca bien foam"],
            "SPL": ["Sensor de PCB detecta que hay placa cuando no la hay","No detecta marcas Power","Colisión placas","Fallo dispensación glue","Marco atascado en parte inferior",
                    "Soldadura defectuosa","Error en sensor de salida"]
        }

        for name, incidences in self.incidencias.items():
            self.create_tab(name, incidences)

    def create_tab(self, name, incidences):
        tab = QWidget()
        self.tabWidget.addTab(tab, name)
        layout = QVBoxLayout(tab)

        title = QLabel(f"Incidencias - {name}")
        title.setFont(QFont('Arial', 16))
        layout.addWidget(title)

        last_incidence_label = QLabel("Última incidencia: Ninguna")
        last_incidence_label.setFont(QFont('Arial', 40))
        layout.addWidget(last_incidence_label)

        self.last_incidence_labels[name] = last_incidence_label

        listWidget = QListWidget()
        for incidence in incidences:
            item = QListWidgetItem(QIcon("app/logo.png"), incidence)
            item.setFont(QFont('Arial', 14))
            listWidget.addItem(item)
        layout.addWidget(listWidget)

        buttonLayout = QHBoxLayout()

        confirmButton = QPushButton("Confirmar")
        confirmButton.setIcon(QIcon("app/logo.png"))
        confirmButton.clicked.connect(lambda: self.on_confirm(name, listWidget.currentItem().text() if listWidget.currentItem() else ""))
        buttonLayout.addWidget(confirmButton)

        layout.addLayout(buttonLayout)

    def on_confirm(self, name, incidence):
        if incidence:
            response = QMessageBox.question(self, "Confirmar Incidencia", f"¿Estás seguro de confirmar la incidencia '{incidence}' en {name}?", QMessageBox.Yes | QMessageBox.No)
            if response == QMessageBox.Yes:
                self.last_incidence = {"bloque": name, "incidencia": incidence, "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
                self.update_last_incidence()
                self.write_to_excel(name, incidence)
        else:
            QMessageBox.warning(self, "Selección Vacía", "Por favor, selecciona una incidencia.")

    
    def create_excel_file(self):
        """Crea el archivo Excel si no existe, incluyendo los encabezados para Fecha y Hora separadamente."""
        if not os.path.exists(self.excel_file):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Incidencias"
            headers = ['Fecha', 'Hora'] + self.blocks  # Agrega 'Fecha' y 'Hora' como las dos primeras columnas
            sheet.append(headers)
            workbook.save(self.excel_file)

    def write_to_excel(self, block, incidence):
        """Escribe las incidencias en el archivo Excel de acuerdo al bloque, separando fecha y hora."""
        workbook = load_workbook(self.excel_file)
        sheet = workbook.active

        # Identificar la columna correcta del bloque
        col_num = self.blocks.index(block) + 3  # Ajuste por las dos nuevas columnas de Fecha y Hora

        # Encuentra la primera fila vacía para registrar la nueva incidencia
        row_num = 2
        while sheet.cell(row=row_num, column=1).value is not None:
            row_num += 1

        # Separar fecha y hora
        current_date = datetime.now().strftime('%Y-%m-%d')
        current_time = datetime.now().strftime('%H:%M:%S')

        # Registrar la fecha y hora en las primeras dos columnas
        sheet.cell(row=row_num, column=1, value=current_date)
        sheet.cell(row=row_num, column=2, value=current_time)

        # Registrar la incidencia en la columna correspondiente al bloque
        sheet.cell(row=row_num, column=col_num, value=incidence)
        workbook.save(self.excel_file)

    def on_network_reply(self, reply):
        if reply.error() == QNetworkReply.NoError:
            QMessageBox.information(self, "Reporte de Incidencias", "La incidencia se ha reportado correctamente.")
        else:
            QMessageBox.warning(self, "Reporte de Incidencias", f"Error al reportar la incidencia: {reply.errorString()}")

    def update_last_incidence(self):
        if self.last_incidence:
            block = self.last_incidence["bloque"]
            incidence = self.last_incidence["incidencia"]
            timestamp = self.last_incidence["timestamp"]
            label_text = f"Última incidencia confirmada: {incidence} a las {timestamp}"
            if block in self.last_incidence_labels:
                self.last_incidence_labels[block].setText(label_text)

    def update_status_bar(self):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.status_bar.showMessage(f"Hora actual: {current_time}")

if __name__ == "__main__":
    app = QApplication(sys.argv)

    login_window = LoginWindow()
    ticket_management = TicketManagement()

    def on_login_successful():
        ticket_management.show()

    login_window.login_successful.connect(on_login_successful)
    login_window.show()

    sys.exit(app.exec_())