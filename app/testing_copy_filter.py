import sys
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QWidget, QPushButton, QMessageBox, QTabWidget, QLabel,
    QListWidget, QListWidgetItem, QStatusBar, QLineEdit, QFileDialog, QHBoxLayout, QTableWidget, QTableWidgetItem, 
    QDialog, QComboBox, QScrollArea
)
from PyQt5.QtCore import QTimer, Qt, pyqtSignal, QRect
from PyQt5.QtGui import QFont, QPixmap, QColor

from collections import Counter, defaultdict

# Clase User para almacenar información de usuario
class User:
    def __init__(self, username, password, blocks):
        self.username = username
        self.password = password
        self.blocks = blocks

# Clase LoginWindow para manejar inicio de sesión
class LoginWindow(QWidget):
    login_successful = pyqtSignal(User)

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Inicio de Sesión")
        self.resize(280, 360)
        self.center_window()
        self.setStyleSheet("background-color: white;")

        # Lista de usuarios
        self.users = [
            User("pideu1", "1111", ["WC47 NACP"]),
            User("pideu2", "1111", ["WC48 P5F"]),
            User("pideu3", "1111", ["WC49 P5H"]),
            User("pideu4", "1111", ["WV50 FILTER"]),
            User("pideu5", "1111", ["SPL"])
        ]

        layout = QVBoxLayout()
        logo_pixmap = QPixmap("app/logo.png")

        logo_label = QLabel()
        logo_label.setPixmap(logo_pixmap.scaled(200, 200, Qt.KeepAspectRatio))
        logo_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(logo_label)

        font = QFont()
        font.setPointSize(12)

        self.username_input = QLineEdit()
        self.username_input.setFont(font)
        self.username_input.setPlaceholderText("Usuario")
        layout.addWidget(self.username_input)

        self.password_input = QLineEdit()
        self.password_input.setFont(font)
        self.password_input.setPlaceholderText("Contraseña")
        self.password_input.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.password_input)

        login_button = QPushButton("Confirmar")
        login_button.clicked.connect(self.login)
        layout.addWidget(login_button)

        self.setLayout(layout)

    def center_window(self):
        screen_rect = QApplication.desktop().availableGeometry()
        x = (screen_rect.width() - self.width()) // 2
        y = (screen_rect.height() - self.height()) // 2
        self.setGeometry(QRect(x, y, self.width(), self.height()))

    def login(self):
        username = self.username_input.text()
        password = self.password_input.text()
        for user in self.users:
            if username == user.username and password == user.password:
                self.login_successful.emit(user)
                self.close()
                return
        QMessageBox.warning(self, "Error de Inicio de Sesión", "Usuario o contraseña incorrectos")

# Clase AdvancedFilterDialog para el filtro avanzado
class AdvancedFilterDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Filtro Avanzado")
        self.setGeometry(300, 300, 800, 600)

        layout = QVBoxLayout()

        range_layout = QHBoxLayout()

        date_layout = QVBoxLayout()
        date_layout.addWidget(QLabel("Desde Fecha:"))
        self.start_date_combo = QComboBox()
        date_layout.addWidget(self.start_date_combo)

        date_layout.addWidget(QLabel("Hasta Fecha:"))
        self.end_date_combo = QComboBox()
        date_layout.addWidget(self.end_date_combo)

        range_layout.addLayout(date_layout)

        time_layout = QVBoxLayout()
        time_layout.addWidget(QLabel("Desde Hora:"))
        self.start_time_combo = QComboBox()
        self.populate_time_combo(self.start_time_combo)
        time_layout.addWidget(self.start_time_combo)

        time_layout.addWidget(QLabel("Hasta Hora:"))
        self.end_time_combo = QComboBox()
        self.populate_time_combo(self.end_time_combo)
        time_layout.addWidget(self.end_time_combo)

        range_layout.addLayout(time_layout)

        layout.addLayout(range_layout)

        block_layout = QHBoxLayout()
        block_layout.addWidget(QLabel("Seleccionar Bloque:"))
        self.block_selector = QComboBox()
        self.block_selector.addItem("Todos")
        self.block_selector.addItems(parent.blocks)
        block_layout.addWidget(self.block_selector)

        layout.addLayout(block_layout)

        filter_button = QPushButton("Aplicar Filtro")
        filter_button.clicked.connect(self.apply_filter)
        layout.addWidget(filter_button)

        self.tab_widget = QTabWidget()
        layout.addWidget(self.tab_widget)

        self.results_tab = QWidget()
        self.results_layout = QVBoxLayout(self.results_tab)
        self.tab_widget.addTab(self.results_tab, "Resultados")

        self.results_table = QTableWidget()
        self.results_table.setColumnCount(3)
        self.results_table.setHorizontalHeaderLabels(["Bloque", "Número de Incidencias", "Incidencia más frecuente"])
        self.results_layout.addWidget(self.results_table)

        self.incidents_table = QTableWidget()
        self.incidents_table.setColumnCount(2)
        self.incidents_table.setHorizontalHeaderLabels(["Incidencia", "Número de Incidencias"])
        self.results_layout.addWidget(self.incidents_table)

        self.graph_button = QPushButton("Ver Gráficos")
        self.graph_button.clicked.connect(self.open_graph_dialog)
        layout.addWidget(self.graph_button)

        self.setLayout(layout)
        self.populate_date_combos()

    def populate_date_combos(self):
        today = datetime.today()
        for i in range(30):
            date = today - timedelta(days=i)
            date_str = date.strftime("%Y-%m-%d")
            self.start_date_combo.addItem(date_str)
            self.end_date_combo.addItem(date_str)

    def populate_time_combo(self, combo):
        for hour in range(24):
            time_str = f"{hour:02}:00"
            combo.addItem(time_str)

    def apply_filter(self):
        start_date = self.start_date_combo.currentText()
        start_time = self.start_time_combo.currentText()
        start_dt = datetime.strptime(f"{start_date} {start_time}:00", "%Y-%m-%d %H:%M:%S")

        end_date = self.end_date_combo.currentText()
        end_time = self.end_time_combo.currentText()
        end_dt = datetime.strptime(f"{end_date} {end_time}:00", "%Y-%m-%d %H:%M:%S")

        selected_block = self.block_selector.currentText()

        results, trends = self.parent().get_filtered_incidents(start_dt, end_dt, selected_block)

        self.results_table.setRowCount(len(results))
        for row, (block, data) in enumerate(results.items()):
            self.results_table.setItem(row, 0, QTableWidgetItem(block))
            self.results_table.setItem(row, 1, QTableWidgetItem(str(data['count'])))
            self.results_table.setItem(row, 2, QTableWidgetItem(data['most_common_incidence']))

        self.update_incidents_table(results)
        self.trends = results

    def update_incidents_table(self, results):
        all_incidents = []
        for block, data in results.items():
            all_incidents.extend(data['incidences'])

        counter = Counter(all_incidents)
        sorted_incidents = counter.most_common()

        self.incidents_table.setRowCount(len(sorted_incidents))
        for row, (incident, count) in enumerate(sorted_incidents):
            self.incidents_table.setItem(row, 0, QTableWidgetItem(incident))
            self.incidents_table.setItem(row, 1, QTableWidgetItem(str(count)))

    def open_graph_dialog(self):
        if hasattr(self, 'trends'):
            self.graph_dialog = GraphDialog(self, data=self.trends)
            self.graph_dialog.exec_()
        else:
            QMessageBox.warning(self, "Error", "Primero aplica el filtro para ver los gráficos.")

# Clase para mostrar detalles de incidencias más relevantes
class TopIncidentsDialog(QDialog):
    def __init__(self, parent=None, incident_details=None):
        super().__init__(parent)
        self.setWindowTitle("Detalles de Incidencias Más Relevantes")
        self.setGeometry(300, 300, 600, 400)

        layout = QVBoxLayout()
        
        scroll_area = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        for block, details in incident_details.items():
            block_label = QLabel(f"Bloque: {block}")
            block_label.setFont(QFont("Arial", 12, QFont.Bold))
            scroll_layout.addWidget(block_label)

            top_incidents = details.most_common(5)
            for incident, count in top_incidents:
                incident_label = QLabel(f"  {incident}: {count}")
                scroll_layout.addWidget(incident_label)

        scroll_area.setWidget(scroll_widget)
        scroll_area.setWidgetResizable(True)
        layout.addWidget(scroll_area)

        close_button = QPushButton("Cerrar")
        close_button.clicked.connect(self.close)
        layout.addWidget(close_button)

        self.setLayout(layout)

# Clase TicketManagement para manejar la gestión de tickets
class TicketManagement(QMainWindow):
    config_file = "config.txt"

    def __init__(self, user):
        super().__init__()
        self.user = user
        self.excel_file = None
        self.blocks = user.blocks
        self.last_incidence_labels = {}
        self.incidencias = {
            "WC47 NACP": ["Etiquetadora", "Fallo en elevador", "No atornilla tapa", "Fallo tolva",
                        "Fallo en paletizador", "No coge placa", "Palet atascado en la curva",
                        "Ascensor no sube", "No pone tornillo", "Fallo tornillo", "AOI no detecta pieza",
                        "No atornilla clips", "Fallo fijador tapa", "Secuencia atornillador",
                        "Fallo atornillador", "Fallo cámara visión"],
            "WC48 P5F": ["Etiquetadora", "AOI (fallo etiqueta)", "AOI (malla)", "Cámara no detecta Pcb", "Cámara no detecta skeleton",
                        "Cámara no detecta foams", "Cámara no detecta busbar", "Cámara no detecta foam derecho", "No detecta presencia power CP",
                        "Tornillo atascado en tolva", "Cámara no detecta Power CP", "Cámara no detecta Top cover", "Detección de sealling mal puesto",
                        "Robot no coge busbar", "Fallo etiqueta", "Power atascado en prensa, cuesta sacar", "No coloca bien el sealling"],
            "WC49 P5H": ["La cámara no detecta Busbar", "La cámara no detecta Top Cover", "Screw K30 no lo detecta puesto", "Atasco tuerca",
                        "Tornillo atascado", "Etiquetadora", "Detección de sealling mal puesto", "No coloca bien el sealling", "Power atascado en prensa, cuesta sacar",
                        "No lee QR"],
            "WV50 FILTER": ["Fallo cámara ferrite", "NOK Soldadura Plástico", "NOK Soldadura metal", "Traza", "NOK Soldad. Plástico+Metal", "Robot no coloca bien filter en palet",
                            "No coloca bien la pcb", "QR desplazado", "Core enganchado", "Robot no coge PCB", "Fallo atornillador", "Pieza enganchada en HV Test", "Cover atascado",
                            "Robot no coloca bien ferrita", "No coloca bien el core", "Fallo Funcional", "Fallo visión core", "Fallo cámara cover", "Repeat funcional", "Fallo cámara QR",
                            "No coloca bien foam"],
            "SPL": ["Sensor de PCB detecta que hay placa cuando no la hay", "No detecta marcas Power", "Colisión placas", "Fallo dispensación glue", "Marco atascado en parte inferior",
                    "Soldadura defectuosa", "Error en sensor de salida"]
        }
        self.incidences_count = {block: 0 for block in self.incidencias.keys()}
        self.initUI()
        self.load_last_excel_file()

    def initUI(self):
        self.setWindowTitle(f"Ticket Management - {self.user.username}")
        self.resize(1200, 800)
        self.center_window_app()

        main_layout = QHBoxLayout()
        central_widget = QWidget()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

        self.tabWidget = QTabWidget(self)
        main_layout.addWidget(self.tabWidget, 2)

        for name in self.blocks:
            self.create_tab(name, self.incidencias[name])

        right_layout = QVBoxLayout()
        select_excel_button = QPushButton("Seleccionar Archivo Excel", self)
        select_excel_button.clicked.connect(self.select_excel_file)
        right_layout.addWidget(select_excel_button)

        self.excel_path_display = QLineEdit(self)
        self.excel_path_display.setReadOnly(True)
        right_layout.addWidget(self.excel_path_display)

        self.table_widget = QTableWidget(self)
        self.table_widget.setFixedSize(600, 300)
        right_layout.addWidget(self.table_widget)

        self.global_incidence_list = QListWidget(self)
        self.global_incidence_list.setFixedSize(600, 200)
        right_layout.addWidget(self.global_incidence_list)

        filter_button = QPushButton("Filtro Avanzado", self)
        filter_button.clicked.connect(self.open_advanced_filter_dialog)
        right_layout.addWidget(filter_button)

        # Sección para incidencias más relevantes
        self.top_incidents_label = QLabel("Incidencias Más Relevantes")
        right_layout.addWidget(self.top_incidents_label)

        self.view_details_button = QPushButton("Ver Detalles")
        self.view_details_button.clicked.connect(self.open_top_incidents_dialog)
        right_layout.addWidget(self.view_details_button)

        right_layout.addStretch()

        right_widget = QWidget()
        right_widget.setLayout(right_layout)
        main_layout.addWidget(right_widget, 1)

        self.status_bar = QStatusBar(self)
        self.setStatusBar(self.status_bar)

        timer = QTimer(self)
        timer.timeout.connect(self.update_status_bar)
        timer.start(1000)

        self.apply_styles()
    
    def center_window_app(self):
        screen_rect = QApplication.desktop().availableGeometry()
        window_width = self.width()
        window_height = self.height()
        x = (screen_rect.width() - window_width) // 2
        y = (screen_rect.height() - window_height) // 2
        self.setGeometry(QRect(x, y, window_width, window_height))

    def open_advanced_filter_dialog(self):
        self.filter_dialog = AdvancedFilterDialog(self)
        self.filter_dialog.exec_()

    def open_top_incidents_dialog(self):
        self.top_incidents_dialog = TopIncidentsDialog(self, incident_details=self.incident_details)
        self.top_incidents_dialog.exec_()

    def get_filtered_incidents(self, start_dt, end_dt, selected_block):
        if not self.excel_file or not os.path.exists(self.excel_file):
            return {}, {}

        workbook = load_workbook(self.excel_file)
        sheet = workbook.active

        filtered_counts = {block: {'count': 0, 'incidences': []} for block in self.incidencias.keys()}
        trends = {block: defaultdict(int) for block in self.incidencias.keys()}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            date_str, time_str, *incidences = row
            row_datetime = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")

            if start_dt <= row_datetime <= end_dt:
                hour_str = row_datetime.strftime("%Y-%m-%d %H:00:00")
                for i, incidence in enumerate(incidences):
                    if incidence != "-":
                        block = list(self.incidencias.keys())[i]
                        if selected_block == "Todos" or selected_block == block:
                            filtered_counts[block]['count'] += 1
                            filtered_counts[block]['incidences'].append(incidence)
                            trends[block][hour_str] += 1

        for block, data in filtered_counts.items():
            if data['incidences']:
                most_common_incidence, _ = Counter(data['incidences']).most_common(1)[0]
                data['most_common_incidence'] = most_common_incidence
            else:
                data['most_common_incidence'] = "N/A"

        if selected_block != "Todos":
            filtered_counts = {selected_block: filtered_counts[selected_block]}
            trends = {selected_block: trends[selected_block]}

        return filtered_counts, trends

    def create_tab(self, name, incidences):
        tab = QWidget()
        self.tabWidget.addTab(tab, name)
        layout = QVBoxLayout(tab)

        title = QLabel(f"Incidencias - {name}")
        layout.addWidget(title)

        last_incidence_label = QLabel("Última incidencia: Ninguna")
        layout.addWidget(last_incidence_label)
        self.last_incidence_labels[name] = last_incidence_label

        list_widget = QListWidget()
        font = QFont("Arial", 12)
        list_widget.setFont(font)
        for incidence in incidences:
            item = QListWidgetItem(incidence)
            list_widget.addItem(item)

        layout.addWidget(list_widget)

        confirm_button = QPushButton("Confirmar Incidencia")
        button_font = QFont("Arial", 10)
        confirm_button.setFont(button_font)
        confirm_button.setFixedSize(500, 40)
        confirm_button.clicked.connect(lambda: self.confirm_incidence(name, list_widget))
        layout.addWidget(confirm_button)

    def confirm_incidence(self, block_name, list_widget):
        current_item = list_widget.currentItem()
        if current_item:
            incidence_text = current_item.text()
            timestamp = datetime.now()
            date_str = timestamp.strftime("%Y-%m-%d")
            time_str = timestamp.strftime("%H:%M:%S")
            self.last_incidence_labels[block_name].setText(f"Incidencia confirmada: {incidence_text} a las {time_str}")
            self.log_incidence_to_excel(block_name, date_str, time_str, incidence_text)

            QMessageBox.information(self, "Confirmación", "Incidencia confirmada.")

            self.update_excel_table()
            self.update_top_incidents()

            self.global_incidence_list.addItem(f"{block_name}: {incidence_text} a las {time_str} del {date_str}")
        else:
            QMessageBox.warning(self, "Ninguna Incidencia Seleccionada", "Selecciona una incidencia para confirmar.")

    def select_excel_file(self):
        file_dialog = QFileDialog()
        file_dialog.setNameFilter("Archivos Excel (*.xlsx)")
        if file_dialog.exec_():
            file_path = file_dialog.selectedFiles()[0]
            self.excel_file = file_path
            self.excel_path_display.setText(file_path)
            with open(self.config_file, "w") as config:
                config.write(file_path)
            self.update_excel_table()
            self.update_top_incidents()

    def create_excel_if_not_exists(self, file_path):
        if not os.path.exists(file_path):
            workbook = Workbook()
            sheet = workbook.active
            headers = ["Fecha", "Hora"] + list(self.incidencias.keys())
            sheet.append(headers)
            workbook.save(file_path)

    def log_incidence_to_excel(self, block_name, date_str, time_str, incidence_text):
        if self.excel_file and os.path.exists(self.excel_file):
            try:
                workbook = load_workbook(self.excel_file)
                sheet = workbook.active

                headers = [cell.value for cell in sheet[1]]
                expected_headers = ["Fecha", "Hora"] + list(self.incidencias.keys())
                if headers != expected_headers:
                    sheet.delete_rows(1, 1)
                    sheet.insert_rows(1)
                    for idx, header in enumerate(expected_headers):
                        sheet.cell(row=1, column=idx + 1, value=header)

                new_row = [date_str, time_str] + ["-"] * len(self.incidencias)
                block_index = list(self.incidencias.keys()).index(block_name) + 2
                new_row[block_index] = incidence_text
                sheet.append(new_row)

                workbook.save(self.excel_file)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al registrar la incidencia en Excel: {e}")

    def update_status_bar(self):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.statusBar().showMessage(f"Fecha y Hora Actual: {current_time}")

    def load_last_excel_file(self):
        if os.path.exists(self.config_file):
            with open(self.config_file, "r") as config:
                file_path = config.read().strip()
                if os.path.exists(file_path):
                    self.excel_file = file_path
                    self.excel_path_display.setText(file_path)
                    self.update_excel_table()
                    self.update_top_incidents()

    def update_excel_table(self):
        if self.excel_file and os.path.exists(self.excel_file):
            try:
                workbook = load_workbook(self.excel_file)
                sheet = workbook.active
                self.table_widget.setRowCount(sheet.max_row - 1)
                self.table_widget.setColumnCount(sheet.max_column)

                headers = [cell.value for cell in sheet[1]]
                self.table_widget.setHorizontalHeaderLabels(headers)

                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                    for col_idx, cell_value in enumerate(row):
                        item = QTableWidgetItem(str(cell_value) if cell_value is not None else "")
                        self.table_widget.setItem(row_idx - 1, col_idx, item)

                # Aplicar estilos
                self.style_excel_table()

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al actualizar la tabla: {e}")

    def style_excel_table(self):
        # Aplicar bordes y colores alternados
        for row in range(self.table_widget.rowCount()):
            for col in range(self.table_widget.columnCount()):
                item = self.table_widget.item(row, col)
                if item:
                    item.setFont(QFont("Arial", 10))
                    if row % 2 == 0:
                        item.setBackground(QColor(240, 240, 240))  # Gris claro para filas pares
                    else:
                        item.setBackground(QColor(255, 255, 255))  # Blanco para filas impares
                    item.setTextAlignment(Qt.AlignCenter)

    def update_top_incidents(self):
        self.incidences_count = {block: 0 for block in self.incidencias.keys()}
        self.incident_details = {block: Counter() for block in self.incidencias.keys()}

        if self.excel_file and os.path.exists(self.excel_file):
            workbook = load_workbook(self.excel_file)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) == len(self.incidencias) + 2:  # Asegurarse de que la fila tiene la longitud esperada
                    for i, block in enumerate(list(self.incidencias.keys()), start=2):
                        if row[i] and row[i] != "-":
                            self.incidences_count[block] += 1
                            self.incident_details[block][row[i]] += 1

    def apply_styles(self):
        title_font = QFont("Arial", 14, QFont.Bold)
        normal_font = QFont("Arial", 12)

        for i in range(self.tabWidget.count()):
            tab = self.tabWidget.widget(i)
            for widget in tab.findChildren(QLabel):
                if "Última incidencia" in widget.text():
                    widget.setFont(normal_font)
                else:
                    widget.setFont(title_font)


if __name__ == "__main__":
    app = QApplication(sys.argv)

    login_window = LoginWindow()

    def on_login_successful(user):
        ticket_management = TicketManagement(user)
        ticket_management.show()

    login_window.login_successful.connect(on_login_successful)
    login_window.show()

    sys.exit(app.exec_())